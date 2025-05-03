#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pandas as pd
import yaml
import os
import logging
import datetime
from fpdf import FPDF
import calendar
import json
import sys
import argparse
import io
import locale
import requests
import time
import xlsxwriter

# 设置stdout为UTF-8编码
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
# 设置stderr为UTF-8编码
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 尝试设置系统默认编码
try:
    # 获取当前系统区域设置
    system_locale = locale.getdefaultlocale()
    logger_prefix = f"系统默认区域: {system_locale}, "
    
    # 尝试设置为简体中文
    locale.setlocale(locale.LC_ALL, 'zh_CN.UTF-8')
except Exception as e:
    # 如果失败，尝试使用系统默认值
    try:
        locale.setlocale(locale.LC_ALL, '')
        logger_prefix = "使用系统默认区域设置, "
    except Exception:
        logger_prefix = "无法设置区域, "

logger = logging.getLogger(__name__)

class BillingManager:
    def __init__(self, config_file='vps_data.yml'):
        """
        初始化账单管理器
        
        Args:
            config_file (str): VPS配置文件路径
        """
        self.config_file = config_file
        self.vps_data = []
        self.nat_total_fee = 0
        self.billing_year = datetime.datetime.now().year
        self.billing_month = datetime.datetime.now().month
        self.exchange_rate_cache = None  # 用于缓存汇率
        
        # 确保字体目录存在
        self.ensure_fonts_directory()
        
        self.load_data()
        
    def ensure_fonts_directory(self):
        """确保fonts目录存在"""
        try:
            # 获取当前脚本所在目录
            script_dir = os.path.dirname(os.path.abspath(__file__))
            fonts_dir = os.path.join(script_dir, 'fonts')
            
            # 如果fonts目录不存在，创建它
            if not os.path.exists(fonts_dir):
                os.makedirs(fonts_dir)
                logger.info(f"已创建字体目录: {fonts_dir}")
            
            # 检查DejaVu字体是否存在
            dejavu_font_path = os.path.join(fonts_dir, 'DejaVuSansCondensed.ttf')
            if not os.path.exists(dejavu_font_path):
                logger.warning(f"DejaVu字体文件不存在: {dejavu_font_path}")
                logger.info("PDF导出将使用默认Helvetica字体")
            
            return fonts_dir
        except Exception as e:
            logger.error(f"检查/创建字体目录时发生错误: {str(e)}")
            return None
    
    def load_data(self):
        """从配置文件加载VPS数据"""
        try:
            with open(self.config_file, 'r', encoding='utf-8') as file:
                data = yaml.safe_load(file)
                
            self.vps_data = data.get('vps_servers', [])
            self.total_bill = data.get('total_bill', 0)
            # 不要从配置文件加载NAT费用，强制每次都重新计算
            self.nat_total_fee = 0
            
            logger.info(f"成功从 {self.config_file} 加载了 {len(self.vps_data)} 台VPS数据")
        except Exception as e:
            logger.error(f"加载VPS数据失败: {str(e)}")
            self.vps_data = []
            self.total_bill = 0
            self.nat_total_fee = 0
    
    def save_data(self):
        """保存VPS数据到配置文件"""
        try:
            # 临时保存NAT费用以便计算总费用
            original_nat_fee = self.nat_total_fee
            
            # 强制NAT费用为0，让它在每次计算时重新计算
            self.nat_total_fee = 0
            
            data = {
                'vps_servers': self.vps_data,
                'total_bill': self.total_bill,
                'nat_fee': self.nat_total_fee  # 保存为0，强制每次启动时重新计算
            }
            
            with open(self.config_file, 'w', encoding='utf-8') as file:
                yaml.dump(data, file, default_flow_style=False, allow_unicode=True)
            
            # 恢复原始的NAT费用值
            self.nat_total_fee = original_nat_fee
                
            logger.info(f"成功保存 {len(self.vps_data)} 台VPS数据到 {self.config_file}")
            return True
        except Exception as e:
            logger.error(f"保存VPS数据失败: {str(e)}")
            return False
    
    def get_vps_by_name(self, vps_name):
        """
        根据名称获取VPS数据
        
        Args:
            vps_name (str): VPS名称
            
        Returns:
            dict: VPS数据，如果不存在则返回None
        """
        for vps in self.vps_data:
            if vps.get('name') == vps_name:
                return vps
        return None
    
    def update_vps(self, vps_name, **kwargs):
        """
        更新VPS数据
        
        Args:
            vps_name (str): VPS名称
            **kwargs: 要更新的字段和值
            
        Returns:
            bool: 是否成功
        """
        vps = self.get_vps_by_name(vps_name)
        if vps:
            for key, value in kwargs.items():
                vps[key] = value
            return self.save_data()
        else:
            logger.error(f"找不到VPS: {vps_name}")
            return False
    
    def add_vps(self, vps_data):
        """
        添加新的VPS
        
        Args:
            vps_data (dict): VPS数据
            
        Returns:
            bool: 是否成功
        """
        vps_name = vps_data.get('name')
        if not vps_name:
            logger.error("VPS数据缺少name字段")
            return False
            
        # 检查是否已存在
        if self.get_vps_by_name(vps_name):
            logger.error(f"VPS已存在: {vps_name}")
            return False
            
        # 添加start_date字段，记录创建时间
        if 'start_date' not in vps_data:
            vps_data['start_date'] = datetime.datetime.now().strftime("%Y/%m/%d")
            
        # 添加purchase_date字段，记录购买时间
        if 'purchase_date' not in vps_data:
            vps_data['purchase_date'] = datetime.datetime.now().strftime("%Y/%m/%d")
            
        self.vps_data.append(vps_data)
        return self.save_data()
    
    def delete_vps(self, vps_name):
        """
        删除VPS
        
        Args:
            vps_name (str): VPS名称
            
        Returns:
            bool: 是否成功
        """
        for i, vps in enumerate(self.vps_data):
            if vps.get('name') == vps_name:
                del self.vps_data[i]
                return self.save_data()
                
        logger.error(f"找不到VPS: {vps_name}")
        return False
    
    def set_vps_status(self, vps_name, status):
        """
        设置VPS状态
        
        Args:
            vps_name (str): VPS名称
            status (str): 新状态，例如"在用"或"销毁"
            
        Returns:
            bool: 是否成功
        """
        vps = self.get_vps_by_name(vps_name)
        if vps:
            vps['status'] = status
            
            # 如果状态是销毁，设置销毁日期
            if status == "销毁":
                today = datetime.datetime.now().strftime("%Y/%m/%d")
                vps['cancel_date'] = today
                
            return self.save_data()
        else:
            logger.error(f"找不到VPS: {vps_name}")
            return False
    
    def get_all_vps(self):
        """
        获取所有VPS数据
        
        Returns:
            list: VPS数据列表
        """
        return self.vps_data
    
    def get_active_vps(self):
        """
        获取所有在用的VPS
        
        Returns:
            list: 在用的VPS数据列表
        """
        return [vps for vps in self.vps_data if vps.get('status') == "在用"]
    
    def get_inactive_vps(self):
        """
        获取所有已销毁的VPS
        
        Returns:
            list: 已销毁的VPS数据列表
        """
        return [vps for vps in self.vps_data if vps.get('status') == "销毁"]
    
    def reset_nat_fee(self):
        """
        重置NAT费用，强制系统在下次请求时重新计算
        同时重置汇率缓存，确保使用最新汇率
        """
        self.nat_total_fee = 0
        self.exchange_rate_cache = None  # 重置汇率缓存
        logger.info("已重置NAT费用计算和汇率缓存")
        return True
    
    def get_exchange_rate(self, year=None, month=None):
        """
        获取美元兑人民币汇率 (1人民币=多少美元)
        通过在线API获取指定年月的汇率，如果无法获取则使用固定汇率
        
        Args:
            year (int, optional): 年份，默认为当前年份
            month (int, optional): 月份，默认为当前月份
        
        Returns:
            float: 汇率 (1人民币=多少美元)
        """
        try:
            # 如果没有指定年月，使用当前年月
            if year is None or month is None:
                today = datetime.datetime.now()
                year = year or today.year
                month = month or today.month
                
            # 当月的汇率使用当前获取的实时汇率
            current_year = datetime.datetime.now().year
            current_month = datetime.datetime.now().month
            is_current_month = (year == current_year and month == current_month)
            
            # 缓存文件路径
            exchange_rate_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exchange_rates')
            if not os.path.exists(exchange_rate_dir):
                os.makedirs(exchange_rate_dir)
                
            # 使用年月作为缓存文件名
            cache_filename = f"exchange_rate_{year}_{month}.json"
            exchange_rate_file = os.path.join(exchange_rate_dir, cache_filename)
            current_time = time.time()
            
            # 如果是当前月份，且缓存文件存在但未过期（24小时内），则直接使用缓存汇率
            if is_current_month and os.path.exists(exchange_rate_file):
                with open(exchange_rate_file, 'r') as f:
                    cache_data = json.load(f)
                    # 当前月汇率缓存24小时
                    if current_time - cache_data['timestamp'] < 86400:  # 24小时 = 86400秒
                        logger.info(f"从缓存获取{year}年{month}月汇率: 1元人民币 = {cache_data['rate']:.4f}美元")
                        return round(cache_data['rate'], 4)  # 保留4位小数
            
            # 如果是历史月份，且缓存文件存在，直接使用缓存（历史汇率不变）
            elif not is_current_month and os.path.exists(exchange_rate_file):
                with open(exchange_rate_file, 'r') as f:
                    cache_data = json.load(f)
                    logger.info(f"从缓存获取历史{year}年{month}月汇率: 1元人民币 = {cache_data['rate']:.4f}美元")
                    return round(cache_data['rate'], 4)
            
            # 对于当前月份，使用外部API获取最新汇率
            if is_current_month:
                url = "https://api.exchangerate-api.com/v4/latest/CNY"
                response = requests.get(url, timeout=10)
                data = response.json()
                rate = data['rates']['USD']
                
                # 缓存当前月份汇率
                with open(exchange_rate_file, 'w') as f:
                    json.dump({
                        'rate': rate,
                        'timestamp': current_time,
                        'year': year,
                        'month': month
                    }, f)
                
                logger.info(f"获取{year}年{month}月实时汇率成功: 1元人民币 = {rate:.4f}美元")
                return round(rate, 4)  # 保留4位小数
            
            # 对于历史月份，尝试获取历史汇率（此处可以集成历史汇率API，如果有的话）
            # 由于大多数免费API不提供历史汇率，这里使用估算值
            # 可以根据实际情况调整或集成付费API
            else:
                # 假设每年有一个大致的汇率估算
                estimated_rates = {
                    2024: 0.1408,  # 约7.1人民币=1美元 (2024年估计值)
                    2023: 0.1429,  # 约7.0人民币=1美元 (2023年估计值)
                    2022: 0.1495,  # 约6.69人民币=1美元 (2022年估计值)
                    2021: 0.1550,  # 约6.45人民币=1美元 (2021年估计值)
                }
                
                # 获取年度估算汇率，如果没有则使用默认值
                rate = estimated_rates.get(year, 0.1385)  # 默认约7.22人民币=1美元
                
                # 缓存历史月份汇率，避免重复计算
                with open(exchange_rate_file, 'w') as f:
                    json.dump({
                        'rate': rate,
                        'timestamp': current_time,
                        'year': year,
                        'month': month,
                        'is_estimated': True
                    }, f)
                
                logger.info(f"使用{year}年{month}月估算汇率: 1元人民币 = {rate:.4f}美元")
                return round(rate, 4)  # 保留4位小数
            
        except Exception as e:
            logger.warning(f"获取{year}年{month}月汇率失败: {str(e)}，使用默认汇率")
            # 使用固定汇率: 1人民币 = 0.1385美元 (约7.22人民币=1美元)
            return 0.1385
    
    def calculate_nat_fee(self, year=None, month=None):
        """
        计算NAT费用 - 按实际使用天数统计
        
        Args:
            year (int, optional): 年份，默认为当前年份
            month (int, optional): 月份，默认为当前月份
        
        Returns:
            float: NAT总费用（美元）
        """
        try:
            # 如果没有指定年月，使用当前设置的账单年月
            if year is None:
                year = self.billing_year
            if month is None:
                month = self.billing_month
            
            # 如果已经计算过NAT费用并且大于0，且是当前月份的计算，直接返回
            current_year = datetime.datetime.now().year
            current_month = datetime.datetime.now().month
            is_current_month = (year == current_year and month == current_month)
            
            if is_current_month and self.nat_total_fee > 0:
                logger.info(f"使用已计算的当前月份NAT费用: {self.nat_total_fee}")
                return self.nat_total_fee
                
            # 获取使用NAT的VPS列表（只要设置了use_nat=True，不管状态如何）
            nat_vps_list = [vps for vps in self.vps_data if vps.get('use_nat', False) is True]
            has_nat_usage = len(nat_vps_list) > 0
            
            # 如果没有设置使用NAT的VPS，直接返回0
            if not has_nat_usage:
                logger.info(f"{year}年{month}月没有VPS设置为使用NAT，NAT费用为0")
                if is_current_month:
                    self.nat_total_fee = 0
                return 0
                
            # 计算使用NAT的VPS的天数总和
            total_nat_days = 0
            logger.info(f"开始计算{year}年{month}月NAT费用，有{len(nat_vps_list)}台VPS设置使用NAT")
            
            active_nat_vps = 0
            for vps in nat_vps_list:
                # 计算该VPS在指定月份的使用时长
                usage_result = self.calculate_usage_period(vps, year, month)
                if isinstance(usage_result, tuple) and len(usage_result) == 4:
                    usage_string, days, hours, minutes = usage_result
                    
                    # 将小时和分钟换算成天的小数部分
                    # 如果超过12小时，按整天计算
                    if hours > 12 or (hours == 12 and minutes > 0):
                        days += 1
                    
                    # 只有使用时长大于0才累加
                    if days > 0:
                        total_nat_days += days
                        active_nat_vps += 1
                        logger.info(f"VPS {vps.get('name')} 在{year}年{month}月使用NAT {days}天")
            
            # 如果指定月份没有实际使用NAT的VPS，返回0
            if active_nat_vps == 0 or total_nat_days == 0:
                logger.info(f"{year}年{month}月没有VPS实际使用NAT，NAT费用为0")
                if is_current_month:
                    self.nat_total_fee = 0
                return 0
                
            logger.info(f"{year}年{month}月NAT总使用天数: {total_nat_days}天")
            
            # 总流量费用（人民币）= 所有VPS的使用天数总和 × 每天1G × 每G1元
            nat_fee_cny = total_nat_days * 1 * 1
            logger.info(f"{year}年{month}月NAT费用(人民币): {nat_fee_cny}元")
            
            # 获取指定月份的人民币兑美元汇率并转换为美元，保留2位小数
            exchange_rate = self.get_exchange_rate(year, month)
            nat_fee_usd = round(nat_fee_cny * exchange_rate, 2)
            cny_per_usd = round(1 / exchange_rate, 2) if exchange_rate > 0 else 0
            logger.info(f"{year}年{month}月NAT费用(美元): {nat_fee_usd}美元 (汇率: 1美元 = {cny_per_usd}人民币)")
            
            # 仅当计算当前月份时才保存计算结果到实例变量
            if is_current_month:
                self.nat_total_fee = nat_fee_usd
                
            return nat_fee_usd
        except Exception as e:
            logger.error(f"计算{year}年{month}月NAT费用失败: {str(e)}")
            if year == self.billing_year and month == self.billing_month:
                self.nat_total_fee = 0
            return 0
    
    def calculate_total_bill(self, year=None, month=None):
        """
        计算总账单金额
        
        Args:
            year (int, optional): 年份，默认为当前年份
            month (int, optional): 月份，默认为当前月份
            
        Returns:
            float: 总账单金额
        """
        # 如果没有指定年月，使用当前设置的账单年月
        if year is None:
            year = self.billing_year
        if month is None:
            month = self.billing_month
            
        # 确保重新计算NAT费用
        self.reset_nat_fee()
        
        # 计算所有VPS的费用总和
        vps_total = sum(float(vps.get('total_price', 0)) for vps in self.vps_data)
        vps_total = round(vps_total, 2)  # 确保VPS总费用精确到2位小数
        
        # 检查是否有设置为使用NAT的VPS，不管其状态如何
        nat_vps_list = [vps for vps in self.vps_data if vps.get('use_nat', False) is True]
        has_nat_vps = len(nat_vps_list) > 0
        
        # 获取NAT费用 - 只有当存在设置为使用NAT的VPS时才计算
        nat_fee = 0
        if has_nat_vps:
            nat_fee = self.calculate_nat_fee(year, month)
            nat_fee = round(nat_fee, 2)  # 确保NAT费用精确到2位小数
        
        # 保存NAT总费用（仅当计算当前月份时）
        current_year = datetime.datetime.now().year
        current_month = datetime.datetime.now().month
        is_current_month = (year == current_year and month == current_month)
        if is_current_month:
            self.nat_total_fee = nat_fee
        
        # 计算总费用
        total = vps_total + nat_fee
        total = round(total, 2)  # 确保总费用精确到2位小数
        
        # 仅当计算当前月份时才更新总账单金额
        if is_current_month:
            self.total_bill = total
        
        logger.info(f"计算{year}年{month}月总账单金额: VPS总费用={vps_total:.2f}, NAT费用={nat_fee:.2f}, 总计={total:.2f}")
        return total
    
    def to_dataframe(self, year=None, month=None):
        """
        将VPS数据转换为DataFrame
        
        Args:
            year (int, optional): 指定年份，默认为当前设置的账单年份
            month (int, optional): 指定月份，默认为当前设置的账单月份
            
        Returns:
            DataFrame: VPS数据
        """
        # 强制重置NAT费用计算
        self.reset_nat_fee()
        
        # 使用指定年月或默认设置的年月
        billing_year = year if year is not None else self.billing_year
        billing_month = month if month is not None else self.billing_month
        
        logger.info(f"生成DataFrame账单数据 - 年份: {billing_year}, 月份: {billing_month}")
        
        columns = ["VPS名称", "国家/地区", "使用状态", "销毁时间", "统计截止时间", "使用时长", "月单价", "总金额"]
        data = []
        
        # 确保使用当前实时时间计算
        current_time = datetime.datetime.now()
        
        # 计算当前月份的结束日期
        next_month_year = billing_year
        next_month = billing_month + 1
        if next_month > 12:
            next_month = 1
            next_month_year += 1
        month_end = datetime.datetime(next_month_year, next_month, 1) - datetime.timedelta(days=1)
        
        # 设置月份名称
        month_names = {
            1: "一月", 2: "二月", 3: "三月", 4: "四月",
            5: "五月", 6: "六月", 7: "七月", 8: "八月",
            9: "九月", 10: "十月", 11: "十一月", 12: "十二月"
        }
        month_name = month_names.get(billing_month, str(billing_month) + "月")
        
        # 遍历VPS数据，计算对应月份的使用时长和费用
        for vps in self.vps_data:
            # 重新计算指定月份的使用时长
            usage_result = self.calculate_usage_period(vps, billing_year, billing_month)
            if isinstance(usage_result, tuple) and len(usage_result) == 4:
                usage_string, days, hours, minutes = usage_result
                
                # 只有使用时长大于0的才添加到账单
                if days > 0 or hours > 0 or minutes > 0:
                    # 实时计算价格 - 使用更精确的计算方法
                    price_per_month = vps.get('price_per_month', 0)
                    total_price = self.calculate_price_with_purchase_date(vps, billing_year, billing_month)
                    total_price = round(total_price, 2)  # 确保总价精确到2位小数
                    
                    # 准备行数据
                    row = [
                        vps.get('name', ''),
                        vps.get('country', ''),
                        vps.get('status', ''),
                        vps.get('cancel_date', '') if vps.get('status') == "销毁" else '',
                        f"{billing_year}年{month_name}",
                        usage_string,
                        vps.get('price_per_month', 0),
                        total_price
                    ]
                    data.append(row)
        
        df = pd.DataFrame(data, columns=columns)
        
        # 获取所有设置为使用NAT的VPS（use_nat属性为True，不管状态如何）
        nat_vps_list = [vps for vps in self.vps_data if vps.get('use_nat', False) is True]
        
        # 只有存在设置为使用NAT的VPS时才计算NAT费用
        nat_fee = 0
        if nat_vps_list:
            # 使用指定年月计算NAT费用
            nat_fee = self.calculate_nat_fee(billing_year, billing_month)
            logger.info(f"在to_dataframe中计算的{billing_year}年{billing_month}月NAT费用: {nat_fee}")
        
        # 添加NAT总费用行（仅当NAT费用大于0时）
        if nat_fee > 0:
            # 获取指定年月的实时汇率
            exchange_rate = self.get_exchange_rate(billing_year, billing_month)
            exchange_rate_display = round(1 / exchange_rate, 2) if exchange_rate > 0 else 0  # 显示为美元兑人民币汇率，更直观
            
            # 添加NAT费用说明
            nat_description = f'NAT费用(按当月实时汇率¥{exchange_rate_display}:$1)'
            nat_row = ['', '', '', '', '', '', nat_description, round(nat_fee, 2)]
            df.loc[len(df)] = nat_row
        
        # 计算当前月份的总金额
        total_bill = df["总金额"].sum()
        total_bill = round(total_bill, 2)  # 确保总费用精确到2位小数
        
        # 添加总计行
        total_row = ['', '', '', '', '', '', '总计', total_bill]
        df.loc[len(df)] = total_row
        
        return df
    
    def save_to_excel(self, output_file='vps_billing.xlsx', year=None, month=None):
        """
        将账单数据保存到Excel文件
        
        Args:
            output_file (str): 输出文件路径
            year (int, optional): 年份，默认为当前设置的账单年份
            month (int, optional): 月份，默认为当前设置的账单月份
            
        Returns:
            bool: 是否成功保存
        """
        # 使用指定年月或默认设置的年月
        billing_year = year if year is not None else self.billing_year
        billing_month = month if month is not None else self.billing_month
        
        # 获取账单数据
        bill_data = self.get_monthly_bill_data(billing_year, billing_month)
        
        try:
            bill_rows = bill_data.get('账单行', [])
            
            if not bill_rows:
                logger.warning(f"没有找到{billing_year}年{billing_month}月的账单数据")
                return False
                
            # 创建Excel工作簿
            workbook = xlsxwriter.Workbook(output_file)
            worksheet = workbook.add_worksheet(f"{billing_year}年{billing_month}月账单")
            
            # 定义表头样式
            header_format = workbook.add_format({
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#D9E1F2',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 11
            })
            
            # 定义数据单元格样式
            cell_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10
            })
            
            # 定义货币格式
            money_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'num_format': '$#,##0.00'
            })
            
            # 定义小计/总计行样式
            total_format = workbook.add_format({
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#E2EFDA',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'num_format': '$#,##0.00'
            })
            
            # 定义销毁VPS的样式（红色字体）
            destroyed_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'color': 'red'
            })
            
            # 定义销毁VPS的货币格式（红色字体）
            destroyed_money_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'num_format': '$#,##0.00',
                'color': 'red'
            })
            
            # 定义使用NAT的VPS样式（紫色字体）
            nat_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'color': 'purple'
            })
            
            # 定义使用NAT的VPS货币格式（紫色字体）
            nat_money_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'num_format': '$#,##0.00',
                'color': 'purple'
            })
            
            # 定义非NAT的VPS样式（蓝色字体）
            non_nat_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'color': 'blue'
            })
            
            # 定义非NAT的VPS货币格式（蓝色字体）
            non_nat_money_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'num_format': '$#,##0.00',
                'color': 'blue'
            })
            
            # 设置列宽
            worksheet.set_column('A:A', 12)  # VPS名称
            worksheet.set_column('B:B', 20)  # IP地址
            worksheet.set_column('C:C', 20)  # 国家/地区
            worksheet.set_column('D:D', 15)  # 是否使用NAT
            worksheet.set_column('E:E', 10)  # 使用状态
            
            # 判断是否需要显示销毁时间列
            has_destroyed_vps = bill_data.get('显示销毁时间列', False)
            
            if has_destroyed_vps:
                worksheet.set_column('F:F', 15)  # 销毁时间
                worksheet.set_column('G:G', 20)  # 使用时长
                worksheet.set_column('H:H', 15)  # 单价/月
                worksheet.set_column('I:I', 15)  # 合计
            else:
                worksheet.set_column('F:F', 20)  # 使用时长
                worksheet.set_column('G:G', 15)  # 单价/月
                worksheet.set_column('H:H', 15)  # 合计
            
            # 写入标题
            title = f"{billing_year}年{billing_month}月账单详情"
            max_col = 'I' if has_destroyed_vps else 'H'
            worksheet.merge_range(f'A1:{max_col}1', title, header_format)
            
            # 写入表头
            headers = ['VPS名称', 'IP地址', '国家/地区', '是否使用NAT', '使用状态']
            if has_destroyed_vps:
                headers.append('销毁时间')
            headers.extend(['使用时长', '单价/月（$）', '合计（$）'])
            
            for col, header in enumerate(headers):
                worksheet.write(1, col, header, header_format)
            
            # 写入账单数据
            row_idx = 2
            for row in bill_rows:
                # 基于是否需要显示销毁时间列，决定列的偏移量
                offset = 0
                
                # 检查当前行的VPS是否为销毁状态且本月销毁
                is_destroyed = row['使用状态'] == '销毁'
                # 检查是否使用NAT
                use_nat = row['是否使用NAT'] == '是'
                
                # 根据条件选择格式
                current_cell_format = cell_format
                current_money_format = money_format
                
                if is_destroyed:
                    # 本月销毁的VPS整行显示为红色
                    current_cell_format = destroyed_format
                    current_money_format = destroyed_money_format
                elif use_nat:
                    # 使用NAT的VPS整行显示为紫色
                    current_cell_format = nat_format
                    current_money_format = nat_money_format
                else:
                    # 没有使用NAT的VPS整行显示为蓝色
                    current_cell_format = non_nat_format
                    current_money_format = non_nat_money_format
                
                worksheet.write(row_idx, 0, row['VPS名称'], current_cell_format)
                worksheet.write(row_idx, 1, row.get('IP地址', ''), current_cell_format)
                worksheet.write(row_idx, 2, row.get('国家/地区', ''), current_cell_format)
                worksheet.write(row_idx, 3, row['是否使用NAT'], current_cell_format)
                worksheet.write(row_idx, 4, row['使用状态'], current_cell_format)
                
                if has_destroyed_vps:
                    worksheet.write(row_idx, 5, row.get('销毁时间', ''), current_cell_format)
                    offset = 1
                
                worksheet.write(row_idx, 5 + offset, row['使用时长'], current_cell_format)
                worksheet.write(row_idx, 6 + offset, row['单价/月（$）'], current_money_format)
                worksheet.write(row_idx, 7 + offset, row['合计（$）'], current_money_format)
                
                row_idx += 1
            
            # 写入NAT费用和总计
            end_col = 8 if has_destroyed_vps else 7
            
            # NAT费用
            nat_fee = bill_data.get('NAT费用', 0)
            worksheet.merge_range(f'A{row_idx + 1}:G{row_idx + 1}', 'NAT费用', total_format)
            worksheet.write(row_idx, end_col, nat_fee, total_format)
            
            # 总计
            total = bill_data.get('月总费用', 0)
            worksheet.merge_range(f'A{row_idx + 2}:G{row_idx + 2}', '总计', total_format)
            worksheet.write(row_idx + 1, end_col, total, total_format)
            
            # 添加统计表格
            stats_start_row = row_idx + 4  # 留一行空白
            
            # 计算NAT和非NAT的VPS数量和金额
            nat_vps_count = 0
            nat_vps_cost = 0
            non_nat_vps_count = 0
            non_nat_vps_cost = 0
            
            for row in bill_rows:
                if row['是否使用NAT'] == '是':
                    nat_vps_count += 1
                    nat_vps_cost += row['合计（$）']
                else:
                    non_nat_vps_count += 1
                    non_nat_vps_cost += row['合计（$）']
            
            # 设置表格标题
            stats_title_format = workbook.add_format({
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#B7DEE8',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 11
            })
            
            worksheet.merge_range(f'A{stats_start_row}:C{stats_start_row}', '账单统计信息', stats_title_format)
            
            # 设置表头样式
            stats_header_format = workbook.add_format({
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#E2EFDA',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10
            })
            
            # 设置数据样式
            stats_data_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10
            })
            
            stats_money_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'font_name': 'Arial',
                'font_size': 10,
                'num_format': '$#,##0.00'
            })
            
            # 写入表头
            worksheet.write(stats_start_row, 0, '类型', stats_header_format)
            worksheet.write(stats_start_row, 1, '数量', stats_header_format)
            worksheet.write(stats_start_row, 2, '金额（$）', stats_header_format)
            
            # 写入NAT VPS数据
            worksheet.write(stats_start_row + 1, 0, '使用NAT的VPS', nat_format)
            worksheet.write(stats_start_row + 1, 1, nat_vps_count, nat_format)
            worksheet.write(stats_start_row + 1, 2, nat_vps_cost, nat_money_format)
            
            # 写入非NAT VPS数据
            worksheet.write(stats_start_row + 2, 0, '未使用NAT的VPS', non_nat_format)
            worksheet.write(stats_start_row + 2, 1, non_nat_vps_count, non_nat_format)
            worksheet.write(stats_start_row + 2, 2, non_nat_vps_cost, non_nat_money_format)
            
            # 写入NAT费用行
            worksheet.write(stats_start_row + 3, 0, 'NAT费用', total_format)
            worksheet.write(stats_start_row + 3, 1, '-', total_format)
            worksheet.write(stats_start_row + 3, 2, nat_fee, total_format)
            
            # 写入总计行
            worksheet.write(stats_start_row + 4, 0, '总计', total_format)
            worksheet.write(stats_start_row + 4, 1, nat_vps_count + non_nat_vps_count, total_format)
            worksheet.write(stats_start_row + 4, 2, total, total_format)
            
            # 设置统计表格区域列宽
            worksheet.set_column('A:A', 20)  # 类型列宽
            worksheet.set_column('B:B', 10)  # 数量列宽
            worksheet.set_column('C:C', 15)  # 金额列宽
            
            # 保存工作簿
            workbook.close()
            
            logger.info(f"账单已保存到: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"保存账单到Excel时出错: {str(e)}", exc_info=True)
            return False
    
    def generate_pdf_bill(self, output_file='vps_billing.pdf'):
        """
        生成PDF格式的账单
        
        Args:
            output_file (str): 输出PDF文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            df = self.to_dataframe()
            
            # 创建PDF对象
            pdf = FPDF()
            pdf.add_page()
            
            # 设置字体 - 使用内置字体而不是尝试加载外部字体
            try:
                # 尝试加载中文字体，如果失败则使用默认字体
                font_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts', 'DejaVuSansCondensed.ttf')
                if os.path.exists(font_path):
                    pdf.add_font('DejaVu', '', font_path, uni=True)
                    pdf.set_font('DejaVu', '', 10)
                    logger.info(f"使用中文字体: {font_path}")
                else:
                    logger.warning(f"中文字体文件不存在: {font_path}，使用Helvetica字体")
                    pdf.set_font('Helvetica', '', 10)
            except Exception as font_error:
                logger.warning(f"设置字体失败: {str(font_error)}，使用默认字体")
                pdf.set_font('Helvetica', '', 10)
            
            # 添加标题
            pdf.cell(200, 10, 'VPS Account Statement', 0, 1, 'C')
            pdf.cell(200, 10, f'Generated: {datetime.datetime.now().strftime("%Y-%m-%d")}', 0, 1, 'C')
            pdf.ln(10)
            
            # 表格头部
            col_widths = [30, 25, 20, 25, 25, 25, 20, 20]
            headers = df.columns
            
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 10, str(header), 1, 0, 'C')
            pdf.ln()
            
            # 表格数据
            for i, row in df.iterrows():
                for j, value in enumerate(row):
                    if i < len(df) - 1 and j == 2 and value == "销毁":
                        # 高亮显示"销毁"状态的行
                        pdf.set_fill_color(255, 255, 0)
                        pdf.cell(col_widths[j], 10, str(value), 1, 0, 'C', 1)
                    elif i == len(df) - 1:
                        # 高亮显示总金额行
                        pdf.set_fill_color(230, 230, 250)
                        pdf.cell(col_widths[j], 10, str(value), 1, 0, 'C', 1)
                    elif j == 6 and 'NAT费用' in str(value):
                        # 高亮显示NAT费用行，使用不同的颜色
                        pdf.set_fill_color(200, 255, 200)  # 浅绿色
                        pdf.cell(col_widths[j], 10, str(value), 1, 0, 'C', 1)
                    else:
                        pdf.cell(col_widths[j], 10, str(value), 1, 0, 'C')
                pdf.ln()
            
            # 保存PDF
            pdf.output(output_file)
            
            logger.info(f"成功生成PDF账单: {output_file}")
            return True
        except Exception as e:
            logger.error(f"生成PDF账单失败: {str(e)}", exc_info=True)
            # 尝试创建一个简单的PDF文件以验证文件写入权限
            try:
                simple_pdf = FPDF()
                simple_pdf.add_page()
                simple_pdf.set_font("Helvetica", size=12)
                simple_pdf.cell(200, 10, txt="Test PDF", ln=True, align="C")
                simple_pdf.output(output_file)
                logger.info(f"成功创建简单的测试PDF: {output_file}")
                return True
            except Exception as simple_e:
                logger.error(f"无法创建简单的测试PDF: {str(simple_e)}", exc_info=True)
                return False
            
    def parse_usage_period(self, usage_period):
        """
        解析使用时长
        
        Args:
            usage_period (str): 使用时长，例如"3个月+15"
            
        Returns:
            tuple: (月数, 天数)
        """
        try:
            if "个月" in usage_period:
                parts = usage_period.split("个月")
                months = int(parts[0])
                days = int(parts[1].strip("+")) if "+" in parts[1] else 0
                return months, days
            else:
                return 0, int(usage_period.strip("天"))
        except Exception as e:
            logger.error(f"解析使用时长失败: {str(e)}")
            return 0, 0
            
    def calculate_usage_period(self, vps, year=None, month=None, now=None):
        """
        实时计算使用时长，按照指定年月的日历实时统计
        从该月1号00:00:00开始计算到当前日期时间或月末，精确到分钟
        
        Args:
            vps (dict): VPS信息
            year (int, optional): 指定年份，默认为当前设置的账单年份
            month (int, optional): 指定月份，默认为当前设置的账单月份
            now (datetime, optional): 当前时间，如果不提供则使用系统当前时间
            
        Returns:
            tuple: (使用时长字符串, 使用天数, 使用小时数, 使用分钟数)
        """
        try:
            # 使用指定年月或默认设置的年月
            billing_year = year if year is not None else self.billing_year
            billing_month = month if month is not None else self.billing_month
            
            # 确保年月是整数类型
            try:
                billing_year = int(billing_year)
                billing_month = int(billing_month)
            except (ValueError, TypeError):
                logger.error(f"无效的年月参数: 年份={billing_year}, 月份={billing_month}")
                return "参数错误", 0, 0, 0
            
            vps_name = vps.get('name', '未知')
            vps_status = vps.get('status', '未知')
            
            logger.info(f"计算VPS使用时长 - VPS: {vps_name}, 状态: {vps_status}, 年月: {billing_year}/{billing_month}")
            
            # 获取当前日期时间，如果提供了now参数则使用它
            current_time = now if now is not None else datetime.datetime.now()
            
            # 计算指定月份的起始时间和结束时间
            month_start = datetime.datetime(billing_year, billing_month, 1, 0, 0, 0)
            
            # 计算指定月份的结束时间（下个月的第一天减去1秒）
            next_month_year = billing_year
            next_month = billing_month + 1
            if next_month > 12:
                next_month = 1
                next_month_year += 1
            month_end = datetime.datetime(next_month_year, next_month, 1, 0, 0, 0) - datetime.timedelta(seconds=1)
            
            # 获取和解析购买日期
            purchase_date_str = vps.get('purchase_date', '')
            if purchase_date_str:
                try:
                    if ' ' in purchase_date_str:  # 如果包含时间
                        purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y/%m/%d %H:%M:%S")
                    else:  # 如果只有日期
                        purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y/%m/%d")
                except ValueError:
                    # 尝试其他可能的日期格式
                    try:
                        purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y-%m-%d")
                    except ValueError:
                        # 尝试更多日期格式
                        try:
                            # 尝试处理可能的格式：年/月/日
                            parts = purchase_date_str.split('/')
                            if len(parts) == 3:
                                year_part = int(parts[0])
                                month_part = int(parts[1])
                                day_part = int(parts[2])
                                purchase_date = datetime.datetime(year_part, month_part, day_part)
                            else:
                                logger.warning(f"无法解析VPS {vps.get('name')} 的购买日期: {purchase_date_str}")
                                purchase_date = None
                        except Exception:
                            logger.warning(f"无法解析VPS {vps.get('name')} 的购买日期: {purchase_date_str}")
                            purchase_date = None
                
                # 如果购买日期在计算月份之后，则使用时长为0
                if purchase_date and (purchase_date.year > billing_year or 
                                    (purchase_date.year == billing_year and purchase_date.month > billing_month)):
                    logger.info(f"VPS {vps_name} 的购买日期 {purchase_date_str} 晚于计算月份 {billing_year}/{billing_month}，使用时长为0")
                    return "0天0小时0分钟", 0, 0, 0
            
            # 检查销毁日期
            cancel_date = None
            if vps.get('status') == "销毁" and (vps.get('expire_date') or vps.get('cancel_date')):
                # 优先使用expire_date，如果没有则尝试使用cancel_date
                cancel_date_str = vps.get('expire_date') or vps.get('cancel_date')
                
                # 解析销毁日期
                try:
                    if ' ' in cancel_date_str:  # 如果包含时间
                        cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y/%m/%d %H:%M:%S")
                    else:  # 如果只有日期
                        cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y/%m/%d")
                        # 设置为当天结束时间
                        cancel_date = cancel_date.replace(hour=23, minute=59, second=59)
                except ValueError:
                    # 尝试其他可能的日期格式
                    try:
                        cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y-%m-%d")
                        cancel_date = cancel_date.replace(hour=23, minute=59, second=59)
                    except ValueError:
                        # 尝试更多日期格式
                        try:
                            # 尝试处理可能的格式：年/月/日
                            parts = cancel_date_str.split('/')
                            if len(parts) == 3:
                                year_part = int(parts[0])
                                month_part = int(parts[1])
                                day_part = int(parts[2])
                                cancel_date = datetime.datetime(year_part, month_part, day_part, 23, 59, 59)
                            else:
                                # 如果无法解析，默认使用当前时间
                                cancel_date = month_end
                                logger.warning(f"无法解析VPS {vps.get('name')} 的销毁日期: {cancel_date_str}，将使用月末")
                        except Exception:
                            # 如果无法解析，默认使用当前时间
                            cancel_date = month_end
                            logger.warning(f"无法解析VPS {vps.get('name')} 的销毁日期: {cancel_date_str}，将使用月末")
                
                # 检查销毁日期与当前计算月份的关系
                # 如果销毁日期在计算月份之后的月份，则在当前月份中状态应该显示为"在用"
                if cancel_date.year > billing_year or (cancel_date.year == billing_year and cancel_date.month > billing_month):
                    logger.info(f"VPS {vps_name} 在计算月份 {billing_year}/{billing_month} 后销毁 ({cancel_date})，当月状态视为'在用'")
                    # 在计算时间时暂时当作非销毁VPS处理
                    temp_status = "在用"
                # 如果销毁日期早于计算月份，则该VPS不应在该月账单中显示
                elif cancel_date.year < billing_year or (cancel_date.year == billing_year and cancel_date.month < billing_month):
                    logger.info(f"VPS {vps_name} 在计算月份 {billing_year}/{billing_month} 前已销毁 ({cancel_date})，不显示在当月账单中")
                    return "0天0小时0分钟", 0, 0, 0
                else:
                    # 销毁日期在当月内，使用实际状态和销毁日期
                    logger.info(f"VPS {vps_name} 在计算月份 {billing_year}/{billing_month} 内销毁 ({cancel_date})，按实际销毁日期计算")
                    temp_status = vps_status
            else:
                temp_status = vps_status
            
            # 获取和解析启用日期
            start_date_str = vps.get('start_date')
            if not start_date_str:
                logger.warning(f"VPS {vps.get('name')} 没有启用日期，无法计算使用时长")
                return "未知", 0, 0, 0
            
            # 解析启用日期
            try:
                if ' ' in start_date_str:  # 如果包含时间
                    start_date = datetime.datetime.strptime(start_date_str, "%Y/%m/%d %H:%M:%S")
                else:  # 如果只有日期
                    start_date = datetime.datetime.strptime(start_date_str, "%Y/%m/%d")
            except ValueError:
                # 尝试其他可能的日期格式
                try:
                    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
                except ValueError:
                    # 尝试更多日期格式
                    try:
                        # 尝试处理可能的格式：年/月/日
                        parts = start_date_str.split('/')
                        if len(parts) == 3:
                            year_part = int(parts[0])
                            month_part = int(parts[1])
                            day_part = int(parts[2])
                            start_date = datetime.datetime(year_part, month_part, day_part)
                        else:
                            logger.error(f"无法解析VPS {vps.get('name')} 的启用日期: {start_date_str}")
                            return "日期错误", 0, 0, 0
                    except Exception:
                        logger.error(f"无法解析VPS {vps.get('name')} 的启用日期: {start_date_str}")
                        return "日期错误", 0, 0, 0
            
            # 最终决定计算的开始时间和结束时间
            # 开始时间：购买日期和月初较晚者
            if purchase_date is not None and purchase_date > month_start:
                start_time = purchase_date
                logger.info(f"VPS {vps_name} 使用购买日期 {purchase_date} 作为开始时间")
            else:
                start_time = month_start
                logger.info(f"VPS {vps_name} 使用月初 {month_start} 作为开始时间")
            
            # 结束时间：销毁日期、当前时间、月末三者最早的
            if temp_status == "销毁" and cancel_date is not None and month_start.year == cancel_date.year and month_start.month == cancel_date.month:
                # 只有当销毁日期在当月时才使用销毁日期作为结束时间
                end_time = cancel_date
                logger.info(f"VPS {vps_name} 使用销毁日期 {cancel_date} 作为结束时间")
            else:
                # 非销毁VPS或者不在当月销毁的VPS
                if now is not None and now.year == billing_year and now.month == billing_month:
                    # 如果是当月，使用当前时间
                    end_time = now
                    logger.info(f"VPS {vps_name} 使用当前时间 {now} 作为结束时间 (实时计算)")
                else:
                    # 如果不是当月，使用月末
                    end_time = month_end
                    logger.info(f"VPS {vps_name} 使用月末 {month_end} 作为结束时间")
            
            # 如果开始时间在月末之后或结束时间在月初之前，则没有使用时间
            if start_time > month_end or end_time < month_start:
                logger.info(f"VPS {vps_name} 在{billing_year}年{billing_month}月没有使用时间，开始时间: {start_time}，结束时间: {end_time}")
                return "0天0小时0分钟", 0, 0, 0
            
            # 确保开始时间和结束时间都在当月范围内
            if start_time < month_start:
                start_time = month_start
                logger.info(f"VPS {vps_name} 的开始时间调整为月初: {month_start}")
            
            if end_time > month_end:
                end_time = month_end
                logger.info(f"VPS {vps_name} 的结束时间调整为月末: {month_end}")
            
            # 计算使用时长
            time_diff = end_time - start_time
            
            # 如果结束时间早于开始时间，返回0使用时长
            if time_diff.total_seconds() < 0:
                logger.warning(f"VPS {vps_name} 的结束时间 {end_time} 早于开始时间 {start_time}，返回0使用时长")
                return "0天0小时0分钟", 0, 0, 0
            
            # 计算天、小时、分钟
            days = time_diff.days
            hours = time_diff.seconds // 3600
            minutes = (time_diff.seconds % 3600) // 60
            
            # 返回使用时长的中文表示
            usage_str = f"{days}天{hours}小时{minutes}分钟"
            logger.info(f"VPS {vps_name} 最终计算的使用时长: {usage_str}")
            return usage_str, days, hours, minutes
            
        except Exception as e:
            logger.error(f"计算使用时长时发生错误: {str(e)}", exc_info=True)
            return "错误", 0, 0, 0
    
    def calculate_price(self, price_per_month, days, hours, minutes, year=None, month=None):
        """
        计算VPS价格，根据购买日期和使用时长灵活计费
        
        Args:
            price_per_month (float): 月单价
            days (int): 天数
            hours (int): 小时数
            minutes (int): 分钟数
            year (int, optional): 指定年份，用于计算月天数
            month (int, optional): 指定月份，用于计算月天数
            
        Returns:
            float: 总价
        """
        try:
            # 使用指定年月或当前月来获取实际天数
            if year is not None and month is not None:
                current_month_days = calendar.monthrange(year, month)[1]
            else:
                # 获取当前月的实际天数
                current_date = datetime.datetime.now()
                current_month_days = calendar.monthrange(current_date.year, current_date.month)[1]
            
            # 计算月的总分钟数 (使用实际月份天数)
            minutes_per_month = current_month_days * 24 * 60
            
            # 计算每分钟价格
            price_per_minute = price_per_month / minutes_per_month
            
            # 计算总使用分钟数
            total_minutes = days * 24 * 60 + hours * 60 + minutes
            
            # 计算总价
            total_price = total_minutes * price_per_minute
            
            return round(total_price, 2)  # 确保总价精确到2位小数
        except Exception as e:
            logger.error(f"计算价格失败: {str(e)}")
            return 0.0

    def calculate_price_with_purchase_date(self, vps, year=None, month=None):
        """
        根据购买日期计算价格，实现灵活的计费方式
        
        Args:
            vps (dict): VPS信息
            year (int, optional): 指定年份，默认为当前设置的账单年份
            month (int, optional): 指定月份，默认为当前设置的账单月份
            
        Returns:
            float: 计算的价格
        """
        try:
            # 获取VPS信息
            price_per_month = float(vps.get('price_per_month', 0))
            if price_per_month == 0:
                return 0.0
                
            # 获取VPS名称，用于日志记录
            vps_name = vps.get('name', '未知')
                
            # 使用指定年月或默认设置的年月
            billing_year = year if year is not None else self.billing_year
            billing_month = month if month is not None else self.billing_month
            
            # 计算当前月的天数 - 使用calendar模块获取准确月份天数
            days_in_month = calendar.monthrange(billing_year, billing_month)[1]
            
            # 获取购买日期
            purchase_date_str = vps.get('purchase_date')
            if not purchase_date_str:
                # 如果没有购买日期，使用启用日期
                purchase_date_str = vps.get('start_date')
                if not purchase_date_str:
                    logger.warning(f"VPS {vps_name} 没有购买日期和启用日期，使用默认按月计费")
                    # 计算使用时长
                    usage_result = self.calculate_usage_period(vps, billing_year, billing_month)
                    if isinstance(usage_result, tuple) and len(usage_result) == 4:
                        _, days, hours, minutes = usage_result
                        return self.calculate_price(price_per_month, days, hours, minutes, billing_year, billing_month)
                    return 0.0
            
            # 解析购买日期
            purchase_date = None
            try:
                if ' ' in purchase_date_str:  # 如果包含时间
                    purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y/%m/%d %H:%M:%S")
                else:  # 如果只有日期
                    purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y/%m/%d")
            except ValueError:
                try:
                    purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y-%m-%d")
                except ValueError:
                    try:
                        # 尝试处理其他格式：年/月/日
                        parts = purchase_date_str.split('/')
                        if len(parts) == 3:
                            y_part = int(parts[0])
                            m_part = int(parts[1])
                            d_part = int(parts[2])
                            purchase_date = datetime.datetime(y_part, m_part, d_part)
                        else:
                            logger.warning(f"无法解析VPS {vps.get('name')} 的购买日期: {purchase_date_str}，使用默认按月计费")
                            # 计算使用时长
                            usage_result = self.calculate_usage_period(vps, billing_year, billing_month)
                            if isinstance(usage_result, tuple) and len(usage_result) == 4:
                                _, days, hours, minutes = usage_result
                                return self.calculate_price(price_per_month, days, hours, minutes, billing_year, billing_month)
                            return 0.0
                    except Exception:
                        logger.warning(f"无法解析VPS {vps.get('name')} 的购买日期: {purchase_date_str}，使用默认按月计费")
                        # 计算使用时长
                        usage_result = self.calculate_usage_period(vps, billing_year, billing_month)
                        if isinstance(usage_result, tuple) and len(usage_result) == 4:
                            _, days, hours, minutes = usage_result
                            return self.calculate_price(price_per_month, days, hours, minutes, billing_year, billing_month)
                        return 0.0
            
            if purchase_date is None:
                logger.warning(f"VPS {vps.get('name')} 的购买日期解析失败，使用默认按月计费")
                usage_result = self.calculate_usage_period(vps, billing_year, billing_month)
                if isinstance(usage_result, tuple) and len(usage_result) == 4:
                    _, days, hours, minutes = usage_result
                    return self.calculate_price(price_per_month, days, hours, minutes, billing_year, billing_month)
                return 0.0
            
            # 获取购买日期的日、月、年
            purchase_day = purchase_date.day
            purchase_month = purchase_date.month
            purchase_year = purchase_date.year
            
            # 获取查询月份的月初和月末
            month_start = datetime.datetime(billing_year, billing_month, 1)
            
            next_month_year = billing_year
            next_month = billing_month + 1
            if next_month > 12:
                next_month = 1
                next_month_year += 1
                
            month_end = datetime.datetime(next_month_year, next_month, 1) - datetime.timedelta(seconds=1)
            
            # 计算VPS的启用日期和可能的销毁日期
            start_date = None
            start_date_str = vps.get('start_date')
            if not start_date_str:
                logger.warning(f"VPS {vps.get('name')} 没有启用日期，使用购买日期")
                start_date = purchase_date
            else:
                try:
                    if ' ' in start_date_str:  # 如果包含时间
                        start_date = datetime.datetime.strptime(start_date_str, "%Y/%m/%d %H:%M:%S")
                    else:  # 如果只有日期
                        start_date = datetime.datetime.strptime(start_date_str, "%Y/%m/%d")
                except ValueError:
                    try:
                        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
                    except ValueError:
                        try:
                            # 尝试处理其他格式：年/月/日
                            parts = start_date_str.split('/')
                            if len(parts) == 3:
                                y_part = int(parts[0])
                                m_part = int(parts[1])
                                d_part = int(parts[2])
                                start_date = datetime.datetime(y_part, m_part, d_part)
                            else:
                                logger.warning(f"无法解析VPS {vps.get('name')} 的启用日期: {start_date_str}，使用购买日期")
                                start_date = purchase_date
                        except Exception:
                            logger.warning(f"无法解析VPS {vps.get('name')} 的启用日期: {start_date_str}，使用购买日期")
                            start_date = purchase_date
            
            # 检查是否已销毁
            cancel_date = None
            if vps.get('status') == "销毁" and vps.get('cancel_date'):
                cancel_date_str = vps.get('cancel_date')
                try:
                    if ' ' in cancel_date_str:  # 如果包含时间
                        cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y/%m/%d %H:%M:%S")
                    else:  # 如果只有日期
                        cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y/%m/%d")
                        cancel_date = cancel_date.replace(hour=23, minute=59, second=59)
                except ValueError:
                    try:
                        cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y-%m-%d")
                        cancel_date = cancel_date.replace(hour=23, minute=59, second=59)
                    except ValueError:
                        try:
                            # 尝试处理其他格式：年/月/日
                            parts = cancel_date_str.split('/')
                            if len(parts) == 3:
                                y_part = int(parts[0])
                                m_part = int(parts[1])
                                d_part = int(parts[2])
                                cancel_date = datetime.datetime(y_part, m_part, d_part, 23, 59, 59)
                            else:
                                logger.warning(f"无法解析VPS {vps.get('name')} 的销毁日期: {cancel_date_str}")
                                cancel_date = None
                        except Exception:
                            logger.warning(f"无法解析VPS {vps.get('name')} 的销毁日期: {cancel_date_str}")
                            cancel_date = None
            
            # 判断VPS在当前计费月的情况
            
            # 情况1: VPS在当前计费月之前就已销毁，不需要计费
            if cancel_date and cancel_date < month_start:
                return 0.0
                
            # 情况2: VPS在当前计费月之后才启用，不需要计费
            if start_date > month_end:
                return 0.0
            
            # 获取当前时间 - 确保使用实际当前时间进行计算
            current_time = datetime.datetime.now()
            
            # 设置计费结束时间
            # 如果计算的是当前月份，使用当前时间（强制实时计算）
            if billing_year == current_time.year and billing_month == current_time.month:
                billing_end = min(current_time, month_end)
                logger.info(f"VPS {vps_name} 正在计算当前月份，使用实时时间 {billing_end} 作为结束时间")
            else:
                # 如果计算历史月份，使用月末时间
                billing_end = month_end
                logger.info(f"VPS {vps_name} 正在计算历史月份，使用月末 {billing_end} 作为结束时间")
                
            # 如果有销毁日期且在当前月内，使用销毁日期作为结束时间
            if cancel_date and month_start <= cancel_date <= month_end:
                billing_end = cancel_date
                logger.info(f"VPS {vps_name} 在当月销毁，使用销毁日期 {billing_end} 作为结束时间")
            
            # 设置计费开始时间
            billing_start = month_start
            
            # 如果启用日期在当前月内，使用启用日期作为开始时间
            if start_date > month_start and start_date <= month_end:
                billing_start = start_date
            
            # 判断VPS是否使用满一个月
            is_full_month = False
            
            # 如果是销毁的VPS并且在当前月内销毁
            if vps.get('status') == "销毁" and cancel_date and month_start.year == cancel_date.year and month_start.month == cancel_date.month:
                # 计算VPS在当月内的使用天数
                first_day_of_month = datetime.datetime(billing_year, billing_month, 1)
                days_used_in_month = (cancel_date - first_day_of_month).days + 1
                
                # 必须严格满足使用整月的条件：
                # 1. 使用天数等于当月的总天数
                # 2. 必须从月初第一天开始使用
                if days_used_in_month == days_in_month and billing_start.day == 1 and billing_start.hour == 0 and billing_start.minute == 0:
                    is_full_month = True
                    logger.info(f"VPS {vps_name} 在{billing_year}年{billing_month}月销毁，但使用满了完整一个月，按整月计费")
                else:
                    logger.info(f"VPS {vps_name} 在{billing_year}年{billing_month}月销毁，使用了{days_used_in_month}天，按实际分钟计费")
            else:
                # 非销毁VPS或不在当月销毁的VPS，计算从billing_start到billing_end的时间
                # 必须严格满足条件：1.从月初开始 2.到月末结束 3.天数等于当月天数
                is_month_start = (billing_start.day == 1 and billing_start.hour == 0 and billing_start.minute == 0)
                is_month_end = (billing_end.day == days_in_month and billing_end.hour == 23 and billing_end.minute == 59)
                days_used = (billing_end - billing_start).days + 1
                
                if is_month_start and is_month_end and days_used == days_in_month:
                    is_full_month = True
                    logger.info(f"VPS {vps_name} 在{billing_year}年{billing_month}月使用了完整一个月，按整月计费")
                else:
                    logger.info(f"VPS {vps_name} 在{billing_year}年{billing_month}月未使用完整一个月，按实际分钟计费")
                
            # 根据是否使用满一个月决定计费方式
            if is_full_month:
                # 使用满一个月，按整月收费
                total_price = price_per_month
                logger.info(f"VPS {vps_name} 在{billing_year}年{billing_month}月计费方式: 整月计费, 收费 ${total_price:.2f}")
            else:
                # 计算使用的天数、小时和分钟，保留精度
                time_diff = billing_end - billing_start
                
                # 获取总秒数
                total_seconds = time_diff.total_seconds()
                
                # 计算总分钟数作为小数（保留小数部分以提高精度）
                total_minutes = total_seconds / 60
                
                # 计算每分钟价格（基于用户设置的月单价）
                minutes_per_month = days_in_month * 24 * 60
                price_per_minute = price_per_month / minutes_per_month
                
                # 计算总价 (分钟数 * 每分钟价格)
                total_price = total_minutes * price_per_minute
                
                # 记录详细计费信息
                days_part = time_diff.days
                hours_part = time_diff.seconds // 3600
                minutes_part = (time_diff.seconds % 3600) // 60
                seconds_part = time_diff.seconds % 60
                
                logger.info(f"VPS {vps_name} 在{billing_year}年{billing_month}月计费方式: 分钟计费")
                logger.info(f"使用时长: {days_part}天{hours_part}小时{minutes_part}分钟{seconds_part}秒 ({total_minutes:.2f}分钟)")
                logger.info(f"月单价: ${price_per_month:.2f}, 分钟单价: ${price_per_minute:.6f}")
                logger.info(f"总计费: {total_minutes:.2f}分钟 × ${price_per_minute:.6f}/分钟 = ${total_price:.2f}")
            
            return round(total_price, 2)
            
        except Exception as e:
            logger.error(f"根据购买日期计算价格时出错: {str(e)}", exc_info=True)
            # 发生错误时，尝试使用旧的计算方法
            usage_result = self.calculate_usage_period(vps, year, month)
            if isinstance(usage_result, tuple) and len(usage_result) == 4:
                _, days, hours, minutes = usage_result
                return self.calculate_price(price_per_month, days, hours, minutes)
            return 0.0
            
    def update_prices(self):
        """
        更新所有VPS的价格
        
        Returns:
            bool: 是否成功
        """
        # 确保使用当前实时时间
        current_time = datetime.datetime.now()
        
        for vps in self.vps_data:
            # 实时计算使用时长，精确到分钟
            usage_result = self.calculate_usage_period(vps, now=current_time)
            if isinstance(usage_result, tuple) and len(usage_result) == 4:
                usage_string, days, hours, minutes = usage_result
                vps['usage_period'] = usage_string
                
                # 使用新的计费方法，确保实时计算
                price_per_month = vps.get('price_per_month', 0)
                if price_per_month:
                    total_price = self.calculate_price_with_purchase_date(vps)
                    vps['total_price'] = round(total_price, 2)  # 确保总价精确到2位小数
            else:
                # 向后兼容老格式
                vps['usage_period'] = usage_result
                
                price_per_month = vps.get('price_per_month', 0)
                
                if vps['usage_period'] and price_per_month:
                    # 使用旧的计算方法
                    total_price = self.calculate_price_legacy(price_per_month, vps['usage_period'])
                    vps['total_price'] = round(total_price, 2)  # 确保总价精确到2位小数
        
        self.calculate_total_bill()
        return self.save_data()
        
    def calculate_price_legacy(self, price_per_month, usage_period):
        """
        旧的价格计算方法，用于兼容
        
        Args:
            price_per_month (float): 月单价
            usage_period (str): 使用时长，例如"3个月+15"
            
        Returns:
            float: 总价
        """
        months, days = self.parse_usage_period(usage_period)
        total = months * price_per_month
        if days > 0:
            # 计算每天的价格
            daily_price = price_per_month / 30.0
            total += days * daily_price
        return round(total, 2)  # 确保总价精确到2位小数
    
    def set_billing_period(self, year, month):
        """
        设置账单计算的年月
        
        Args:
            year (int): 年份
            month (int): 月份
            
        Returns:
            bool: 是否成功设置
        """
        try:
            # 验证年月有效性
            if not (2000 <= year <= 2100 and 1 <= month <= 12):
                logger.error(f"无效的年月: {year}/{month}")
                return False
                
            self.billing_year = year
            self.billing_month = month
            logger.info(f"已设置账单计算年月为: {year}/{month}")
            return True
        except Exception as e:
            logger.error(f"设置账单年月时出错: {str(e)}")
            return False
    
    def get_billing_period(self):
        """
        获取当前账单计算的年月
        
        Returns:
            tuple: (年份, 月份)
        """
        return (self.billing_year, self.billing_month)
    
    def generate_monthly_bill_table(self, start_year=2024, end_year=None, end_month=None):
        """
        生成月账单统计表，包含每月VPS使用情况和费用明细
        
        Args:
            start_year (int): 起始年份
            end_year (int, optional): 结束年份，默认为当前年份
            end_month (int, optional): 结束月份，默认为当前月份
            
        Returns:
            tuple: (summary_df, bill_data) 汇总DataFrame和详细账单数据列表
        """
        try:
            # 如果未指定结束年月，使用当前年月
            now = datetime.datetime.now()
            if end_year is None:
                end_year = now.year
            if end_month is None:
                end_month = now.month
                
            logger.info(f"生成月账单统计表 - 起始年份: {start_year}, 结束年月: {end_year}/{end_month}")
            
            # 获取当前实时时间用于计算
            current_time = datetime.datetime.now()
            
            # 用于存储所有月份的账单数据
            bill_data = []
            
            # 遍历每个月份，生成账单数据
            for year in range(start_year, end_year + 1):
                # 确定月份范围
                month_start = 1
                month_end = 12
                
                if year == start_year:
                    month_start = 1  # 起始年份从1月开始
                    
                if year == end_year:
                    month_end = end_month  # 结束年份到指定月份结束
                
                for month in range(month_start, month_end + 1):
                    logger.info(f"正在生成 {year}年{month}月 账单数据")
                    
                    # 获取所有VPS
                    vps_list = self.vps_data
                    
                    # 设置当前处理的账单周期
                    self.set_billing_period(year, month)
                    
                    # 初始化月数据
                    month_data = []
                    has_destroyed_vps_this_month = False
                    
                    # 检查是否有在当月销毁的VPS
                    for vps in vps_list:
                        if vps.get('status') == "销毁":
                            cancel_date_str = vps.get('cancel_date', '')
                            if cancel_date_str:
                                try:
                                    # 尝试解析yyyy/mm/dd格式
                                    parts = cancel_date_str.split('/')
                                    if len(parts) == 3:
                                        cancel_year = int(parts[0])
                                        cancel_month = int(parts[1])
                                        
                                        if cancel_year == year and cancel_month == month:
                                            has_destroyed_vps_this_month = True
                                            break
                                except:
                                    pass
                    
                    # 遍历每个VPS，计算当月使用情况
                    for vps in vps_list:
                        vps_name = vps.get('name', '未命名')
                        use_nat = vps.get('use_nat', False)
                        
                        # 保存原始状态，用于后续判断
                        original_status = vps.get('status', '')
                        
                        # 如果是销毁状态，需要检查销毁日期是否在查询月份
                        cancel_date = None
                        if original_status == "销毁":
                            cancel_date_str = vps.get('cancel_date', '')
                            try:
                                # 尝试解析yyyy-mm-dd格式
                                cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y/%m/%d")
                                cancel_date = cancel_date.replace(hour=23, minute=59, second=59)
                            except ValueError:
                                try:
                                    # 尝试解析yyyy-mm-dd格式
                                    cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y-%m-%d")
                                    cancel_date = cancel_date.replace(hour=23, minute=59, second=59)
                                except ValueError:
                                    try:
                                        parts = cancel_date_str.split('/')
                                        if len(parts) == 3:
                                            y_part = int(parts[0])
                                            m_part = int(parts[1])
                                            d_part = int(parts[2])
                                            cancel_date = datetime.datetime(y_part, m_part, d_part, 23, 59, 59)
                                        else:
                                            cancel_date = None
                                    except Exception:
                                        cancel_date = None
                            
                            # 如果销毁日期在查询月份之前，则跳过该VPS
                            if cancel_date and (cancel_date.year < year or (cancel_date.year == year and cancel_date.month < month)):
                                logger.info(f"VPS {vps_name} 在 {cancel_date.year}/{cancel_date.month}/{cancel_date.day} 销毁，早于查询月份 {year}/{month}，跳过显示")
                                continue
                            
                            # 如果销毁日期在查询月份之后，则该月显示为"在用"
                            if cancel_date and (cancel_date.year > year or (cancel_date.year == year and cancel_date.month > month)):
                                # 创建临时副本以修改状态
                                vps = vps.copy()
                                vps['status'] = "在用"
                                logger.info(f"VPS {vps_name} 在 {cancel_date.year}/{cancel_date.month}/{cancel_date.day} 销毁，晚于查询月份 {year}/{month}，当月显示为'在用'")
                        
                        # 获取购买日期
                        purchase_date_str = vps.get('purchase_date', '')
                        if not purchase_date_str:
                            purchase_date_str = vps.get('start_date', '')
                        
                        # 如果有启用日期，检查是否在查询月份之后，如果是则跳过
                        if purchase_date_str:
                            purchase_date = None
                            try:
                                if ' ' in purchase_date_str:
                                    purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y/%m/%d %H:%M:%S")
                                else:
                                    purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y/%m/%d")
                            except ValueError:
                                try:
                                    purchase_date = datetime.datetime.strptime(purchase_date_str, "%Y-%m-%d")
                                except ValueError:
                                    try:
                                        parts = purchase_date_str.split('/')
                                        if len(parts) == 3:
                                            y_part = int(parts[0])
                                            m_part = int(parts[1])
                                            d_part = int(parts[2])
                                            purchase_date = datetime.datetime(y_part, m_part, d_part)
                                    except:
                                        purchase_date = None
                            
                            # 如果购买日期晚于查询月份的月末，则跳过
                            if purchase_date:
                                month_end_date = datetime.datetime(year, month, calendar.monthrange(year, month)[1], 23, 59, 59)
                                if purchase_date > month_end_date:
                                    logger.info(f"VPS {vps_name} 的购买/启用日期 {purchase_date_str} 晚于查询月份 {year}/{month} 的月末，跳过显示")
                                    continue
                            
                        # 获取销毁日期（仅用于显示，使用原始状态）
                        display_cancel_date_str = ''
                        # 只有当VPS在当月销毁时才设置销毁时间显示
                        if original_status == "销毁" and cancel_date and cancel_date.year == year and cancel_date.month == month:
                            display_cancel_date_str = vps.get('cancel_date', '')
                        
                        # 计算使用时长，使用实时时间
                        usage_result = self.calculate_usage_period(vps, year, month, now=current_time)
                        usage_str = ''
                        days = 0
                        hours = 0
                        minutes = 0
                        
                        if isinstance(usage_result, tuple) and len(usage_result) == 4:
                            usage_str, days, hours, minutes = usage_result
                            
                        # 如果使用时长为0，则跳过该VPS
                        if days == 0 and hours == 0 and minutes == 0:
                            continue
                            
                        # 计算价格，使用更精确的方法
                        price_per_month = float(vps.get('price_per_month', 0))
                        total_price = self.calculate_price_with_purchase_date(vps, year, month)
                        total_price = round(total_price, 2)  # 确保总价精确到2位小数
                        
                        # 决定在该月账单中显示的状态
                        # 如果是销毁状态，仅在销毁当月显示销毁状态，其他情况显示为"在用"
                        if original_status == "销毁":
                            if cancel_date and cancel_date.year == year and cancel_date.month == month:
                                display_status = "销毁"
                            else:
                                display_status = "在用"
                        else:
                            display_status = vps.get('status')
                        
                        # 创建行数据
                        row_data = {
                            'VPS名称': vps_name,
                            '国家/地区': vps.get('country', ''),
                            '使用状态': display_status,
                            '使用时长': usage_str,
                            '单价/月（$）': price_per_month,
                            '合计（$）': total_price,
                            '是否使用NAT': '是' if use_nat else '否',
                            '购买日期': purchase_date_str,
                            '销毁时间': display_cancel_date_str,
                            'raw_value': vps  # 用于调试，JSON输出时会忽略这个字段
                        }
                        
                        # 只有在当月有VPS销毁时才添加销毁时间列
                        if has_destroyed_vps_this_month:
                            row_data['销毁时间'] = display_cancel_date_str
                            
                        # 添加到月数据
                        month_data.append(row_data)
                    
                    # 计算NAT费用 - 修改为使用指定年月的汇率
                    nat_fee = self.calculate_nat_fee(year, month)
                    nat_fee = round(nat_fee, 2)  # 确保NAT费用精确到2位小数
                    
                    # 计算月总费用
                    month_total = sum(item['合计（$）'] for item in month_data)
                    if nat_fee > 0:
                        month_total += nat_fee
                    month_total = round(month_total, 2)  # 确保月总费用精确到2位小数
                    
                    # 添加到账单数据
                    if month_data:  # 只有当月有数据时才添加
                        bill_data.append({
                            '年份': year,
                            '月份': month,
                            '账单日期': f"{year}/{month}/1",
                            'VPS数量': len(month_data),
                            '月总费用': month_total,
                            'NAT费用': nat_fee,
                            '详细数据': month_data,
                            '显示销毁时间列': has_destroyed_vps_this_month  # 添加标记，指示是否显示销毁时间列
                        })
            
            # 创建DataFrame
            columns = [
                '年份', '月份', '账单日期', 'VPS数量', 'NAT费用', '月总费用'
            ]
            
            summary_data = []
            for bill in bill_data:
                row = [
                    bill['年份'],
                    bill['月份'],
                    bill['账单日期'],
                    bill['VPS数量'],
                    bill['NAT费用'],
                    bill['月总费用']
                ]
                summary_data.append(row)
            
            summary_df = pd.DataFrame(summary_data, columns=columns)
            
            return summary_df, bill_data
        
        except Exception as e:
            logger.error(f"生成月账单表格时出错: {str(e)}", exc_info=True)
            return pd.DataFrame(), []
    
    def get_monthly_bill_data(self, year, month):
        """
        获取指定月份的账单数据
        
        Args:
            year (int): 年份
            month (int): 月份
            
        Returns:
            dict: 账单数据
        """
        try:
            # 确保输入的年月是有效的
            year = int(year)
            month = int(month)
            
            if month < 1 or month > 12:
                raise ValueError(f"无效的月份: {month}，月份必须在1-12之间")
                
            # 用于返回的账单数据结构
            bill_data = {
                '年份': year,
                '月份': month,
                '账单日期': f"{year}/{month}/1",
                'VPS数量': 0,
                'NAT费用': 0,
                '月总费用': 0,
                '账单行': []
            }
            
            # 设置月份名称
            month_names = {
                1: "一月", 2: "二月", 3: "三月", 4: "四月",
                5: "五月", 6: "六月", 7: "七月", 8: "八月",
                9: "九月", 10: "十月", 11: "十一月", 12: "十二月"
            }
            month_name = month_names.get(month, str(month) + "月")
            
            # 强制重置NAT费用
            self.reset_nat_fee()
            
            # 设置临时的账单周期
            self.set_billing_period(year, month)
            
            # 获取所有VPS数据
            vps_list = self.get_all_vps()
            active_vps_count = 0
            
            # 获取当前实时时间用于计算
            current_time = datetime.datetime.now()
            
            # 计算每个VPS在指定月份的使用情况
            for vps in vps_list:
                # 计算使用时长，确保使用实时时间
                usage_result = self.calculate_usage_period(vps, year, month, now=current_time)
                
                if isinstance(usage_result, tuple) and len(usage_result) == 4:
                    usage_string, days, hours, minutes = usage_result
                    
                    # 只有使用时长大于0的才添加到账单
                    if days > 0 or hours > 0 or minutes > 0:
                        active_vps_count += 1
                        
                        # 计算VPS价格，使用更精确的方法
                        price_per_month = vps.get('price_per_month', 0)
                        total_price = self.calculate_price_with_purchase_date(vps, year, month)
                        total_price = round(total_price, 2)  # 确保价格精确到2位小数
                        
                        # 获取原始状态和销毁日期
                        original_status = vps.get('status', '')
                        cancel_date_str = vps.get('cancel_date', '')
                        display_status = original_status
                        display_cancel_date = ''
                        
                        # 如果是销毁状态，需要处理销毁日期和状态显示逻辑
                        if original_status == "销毁" and cancel_date_str:
                            cancel_date = None
                            try:
                                # 尝试解析销毁日期
                                try:
                                    cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y/%m/%d")
                                except ValueError:
                                    try:
                                        cancel_date = datetime.datetime.strptime(cancel_date_str, "%Y-%m-%d")
                                    except ValueError:
                                        try:
                                            parts = cancel_date_str.split('/')
                                            if len(parts) == 3:
                                                y_part = int(parts[0])
                                                m_part = int(parts[1])
                                                d_part = int(parts[2])
                                                cancel_date = datetime.datetime(y_part, m_part, d_part)
                                        except Exception:
                                            pass
                                
                                # 根据销毁日期设置显示状态和时间
                                if cancel_date:
                                    # 只在销毁当月显示"销毁"状态和销毁时间
                                    if cancel_date.year == year and cancel_date.month == month:
                                        display_status = "销毁"
                                        display_cancel_date = cancel_date_str
                                    else:
                                        # 在销毁之前的月份显示为"在用"，不显示销毁时间
                                        display_status = "在用"
                                        display_cancel_date = ''
                                    
                                    # 如果销毁日期在查询月份之前，则应该跳过此VPS
                                    if cancel_date.year < year or (cancel_date.year == year and cancel_date.month < month):
                                        logger.info(f"VPS {vps.get('name', '未命名')} 在 {cancel_date_str} 销毁，早于查询月份 {year}/{month}，跳过显示")
                                        continue
                            except Exception as e:
                                logger.error(f"解析销毁日期出错: {str(e)}")
                        
                        # 添加到账单行
                        bill_row = {
                            'VPS名称': vps.get('name', '未命名'),
                            'IP地址': vps.get('ip_address', ''),
                            '国家/地区': vps.get('country', ''),
                            '使用状态': display_status,
                            '销毁时间': display_cancel_date,
                            '统计截止时间': f"{year}年{month_name}",
                            '使用时长': usage_string,
                            '月单价': vps.get('price_per_month', 0),
                            '总金额': total_price,
                            '是否使用NAT': '是' if vps.get('use_nat', False) else '否',
                            '单价/月（$）': vps.get('price_per_month', 0),
                            '合计（$）': total_price
                        }
                        
                        bill_data['账单行'].append(bill_row)
            
            # 计算当月NAT费用，使用指定年月的汇率
            nat_fee = self.calculate_nat_fee(year, month)
            nat_fee = round(nat_fee, 2)  # 确保NAT费用精确到2位小数
            
            # 计算总费用
            vps_total = sum(float(row.get('总金额', 0)) for row in bill_data['账单行'])
            vps_total = round(vps_total, 2)  # 确保VPS总费用精确到2位小数
            
            total_bill = vps_total + nat_fee
            total_bill = round(total_bill, 2)  # 确保总费用精确到2位小数
            
            # 更新账单数据
            bill_data['VPS数量'] = active_vps_count
            bill_data['NAT费用'] = nat_fee
            bill_data['月总费用'] = total_bill
            
            # 如果NAT费用大于0，添加NAT使用详情
            if nat_fee > 0:
                # 计算使用NAT的VPS数量和总天数
                nat_vps_list = [row for row in bill_data['账单行'] if row.get('是否使用NAT') == '是']
                active_nat_vps = len(nat_vps_list)
                
                # 计算NAT总使用天数
                total_nat_days = 0
                for vps in nat_vps_list:
                    usage_str = vps.get('使用时长', '')
                    if '天' in usage_str:
                        try:
                            days_part = usage_str.split('天')[0]
                            days = int(days_part)
                            
                            # 如果有小时部分，且超过12小时，天数+1
                            if '小时' in usage_str:
                                hours_part = usage_str.split('天')[1].split('小时')[0].strip()
                                if hours_part:
                                    hours = int(hours_part)
                                    if hours > 12:
                                        days += 1
                                        
                            total_nat_days += days
                        except (ValueError, IndexError) as e:
                            logger.warning(f"解析使用时长出错: {usage_str}, {str(e)}")
                
                # 获取指定月份的汇率
                exchange_rate = self.get_exchange_rate(year, month)
                exchange_rate_display = round(1 / exchange_rate, 2) if exchange_rate > 0 else 0
                
                # 添加NAT费用详情到账单数据
                bill_data['NAT详情'] = {
                    'NAT使用VPS数': active_nat_vps,
                    'NAT总天数': total_nat_days,
                    '单价': '¥1/G/天',
                    '汇率': f'¥{exchange_rate_display}:$1',
                    '费用说明': f'{active_nat_vps}台VPS共{total_nat_days}天×1G/天×¥1/G÷当月汇率¥{exchange_rate_display}:$1'
                }
            
            return bill_data
            
        except Exception as e:
            logger.error(f"获取{year}年{month}月账单数据失败: {str(e)}")
            # 返回空数据结构
            return {
                '年份': year,
                '月份': month,
                '账单日期': f"{year}/{month}/1",
                'VPS数量': 0,
                'NAT费用': 0,
                '月总费用': 0,
                '账单行': [],
                '错误': str(e)
            }
    
    def save_monthly_billing_to_excel(self, output_file='月账单统计.xlsx', start_year=2024, specific_year=None, specific_month=None):
        """
        将月账单统计保存到Excel
        
        Args:
            output_file (str): 输出文件路径
            start_year (int): 起始年份
            specific_year (int, optional): 指定年份，如果提供则只输出该年份的数据
            specific_month (int, optional): 指定月份，如果提供则只输出指定年月的数据
            
        Returns:
            bool: 是否成功
        """
        try:
            # 强制重置NAT费用计算
            self.reset_nat_fee()
            
            # 如果指定了年月，则只导出该月的数据
            if specific_year is not None and specific_month is not None:
                # 获取指定月份的账单数据
                bill_data = self.get_monthly_bill_data(specific_year, specific_month)
                
                # 确保目录存在
                output_dir = os.path.dirname(output_file)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir)
                
                # 创建Excel工作簿
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    # 设置月份名称
                    month_names = {
                        1: "一月", 2: "二月", 3: "三月", 4: "四月",
                        5: "五月", 6: "六月", 7: "七月", 8: "八月",
                        9: "九月", 10: "十月", 11: "十一月", 12: "十二月"
                    }
                    month_name = month_names.get(specific_month, str(specific_month) + "月")
                    
                    # 创建工作表名
                    sheet_name = f"{specific_year}年{month_name}"
                    
                    # 确保工作表名称不超过31个字符（Excel限制）
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]
                    
                    # 创建表格
                    columns = ["VPS名称", "IP地址", "国家/地区", "是否使用NAT", "使用状态", "使用时长", "单价", "合计（$）"]
                    
                    # 创建DataFrame
                    month_df = pd.DataFrame(columns=columns)
                    
                    # 添加数据行
                    for row in bill_data.get('账单行', []):
                        month_df = pd.concat([month_df, pd.DataFrame([{
                            'VPS名称': row.get('VPS名称', ''),
                            'IP地址': row.get('IP地址', ''),
                            '国家/地区': row.get('国家/地区', ''),
                            '是否使用NAT': row.get('是否使用NAT', '否'),
                            '使用状态': row.get('使用状态', ''),
                            '使用时长': row.get('使用时长', ''),
                            '单价': row.get('月单价', 0),
                            '合计（$）': row.get('总金额', 0)
                        }])], ignore_index=True)
                    
                    # 添加NAT费用和总计行
                    if bill_data['NAT费用'] > 0:
                        # 获取NAT详情
                        nat_details = bill_data.get('NAT详情', {})
                        nat_vps_count = nat_details.get('NAT使用VPS数', 0)
                        total_nat_days = nat_details.get('NAT总天数', 0)
                        
                        # 获取当月汇率
                        exchange_rate = self.get_exchange_rate(specific_year, specific_month)
                        exchange_rate_display = round(1 / exchange_rate, 2) if exchange_rate > 0 else 0
                        
                        nat_row = pd.Series({
                            'VPS名称': f'NAT费用({nat_vps_count}台VPS共{total_nat_days}天×1G/天×¥1/G÷当月汇率¥{exchange_rate_display}:$1)',
                            '合计（$）': bill_data['NAT费用']
                        })
                        month_df = pd.concat([month_df, pd.DataFrame([nat_row])], ignore_index=True)
                    
                    # 添加总计行
                    total_row = pd.Series({
                        'VPS名称': '总计',
                        '合计（$）': bill_data['月总费用']
                    })
                    month_df = pd.concat([month_df, pd.DataFrame([total_row])], ignore_index=True)
                    
                    # 保存到Excel
                    month_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 获取工作表
                    worksheet = writer.sheets[sheet_name]
                    
                    # 导入样式
                    from openpyxl.styles import PatternFill, Alignment, Font, Color, Border, Side
                    
                    # 设置表头样式
                    header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
                    header_font = Font(name='微软雅黑', size=11, bold=True, color='000000')
                    
                    # 设置单元格对齐方式
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    left_alignment = Alignment(horizontal='left', vertical='center')
                    right_alignment = Alignment(horizontal='right', vertical='center')
                    
                    # 应用表头样式
                    for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                    
                    # 设置列宽
                    worksheet.column_dimensions['A'].width = 20  # VPS名称
                    worksheet.column_dimensions['B'].width = 20  # IP地址
                    worksheet.column_dimensions['C'].width = 20  # 国家/地区
                    worksheet.column_dimensions['D'].width = 15  # 是否使用NAT
                    worksheet.column_dimensions['E'].width = 10  # 使用状态
                    worksheet.column_dimensions['F'].width = 20  # 使用时长
                    worksheet.column_dimensions['G'].width = 10  # 单价
                    worksheet.column_dimensions['H'].width = 15  # 合计
                    
                    # 设置数据行样式
                    for row in range(2, worksheet.max_row + 1):
                        # 设置每一行的样式
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            
                            # VPS名称列左对齐
                            if col == 1:
                                cell.alignment = left_alignment
                            # 金额列右对齐
                            elif col >= 4:
                                cell.alignment = right_alignment
                            else:
                                cell.alignment = center_alignment
                            
                            # 根据不同状态设置不同颜色
                            if row < worksheet.max_row - 1:  # 排除NAT费用和总计行
                                # 获取第B列的使用状态值
                                status_cell = worksheet.cell(row=row, column=2)
                                status = status_cell.value
                                
                                # 获取该行是否使用NAT (如果有该数据)
                                use_nat = None
                                if '是否使用NAT' in month_df.columns:
                                    nat_col = month_df.columns.get_loc('是否使用NAT') + 1
                                    nat_cell = worksheet.cell(row=row, column=nat_col)
                                    use_nat = nat_cell.value
                                
                                # 设置颜色
                                if status == '销毁':
                                    # 销毁的VPS设为红色
                                    cell.font = Font(name='微软雅黑', size=10, color='FF0000')
                                elif use_nat == '是':
                                    # 使用NAT的VPS设为紫色
                                    cell.font = Font(name='微软雅黑', size=10, color='800080')
                                elif use_nat == '否':
                                    # 不使用NAT的VPS设为蓝色
                                    cell.font = Font(name='微软雅黑', size=10, color='0000FF')
                            
                            # 最后两行（NAT费用和总计）使用特殊样式
                            if row >= worksheet.max_row - 1:
                                cell.font = Font(name='微软雅黑', size=11, bold=True, color='000000')
                                
                                # 为NAT费用和总计行的第一列设置特殊格式，合并前7列
                                if col == 1:  # 第一列
                                    # 找出合计列的索引
                                    total_col = None
                                    for c in range(1, worksheet.max_column + 1):
                                        if worksheet.cell(row=1, column=c).value == '合计（$）':
                                            total_col = c
                                            break
                                    
                                    # 如果找到了合计列，合并从第一列到合计列前一列的单元格
                                    if total_col and total_col > 1:
                                        # 获取当前行的名称（"NAT费用"或"总计"）
                                        row_name = worksheet.cell(row=row, column=col).value
                                        
                                        # 合并单元格
                                        merge_range = f"{chr(64 + col)}{row}:{chr(64 + total_col - 1)}{row}"
                                        worksheet.merge_cells(merge_range)
                                        
                                        # 设置合并后单元格的值
                                        worksheet.cell(row=row, column=col).value = row_name
                    
                    # 添加统计表格
                    stats_start_row = worksheet.max_row + 2  # 留一行空白
                    
                    # 计算NAT和非NAT的VPS数量和金额
                    nat_vps_rows = [row for row in bill_data.get('账单行', []) if row.get('是否使用NAT') == '是']
                    non_nat_vps_rows = [row for row in bill_data.get('账单行', []) if row.get('是否使用NAT') != '是']
                    
                    nat_vps_count = len(nat_vps_rows)
                    nat_vps_cost = sum(row.get('总金额', 0) for row in nat_vps_rows)
                    
                    non_nat_vps_count = len(non_nat_vps_rows)
                    non_nat_vps_cost = sum(row.get('总金额', 0) for row in non_nat_vps_rows)
                    
                    nat_fee = bill_data.get('NAT费用', 0)
                    total_amount = bill_data.get('月总费用', 0)
                    
                    # 创建统计表
                    # 设置表格标题
                    title_cell = worksheet.cell(row=stats_start_row, column=1)
                    title_cell.value = '账单统计信息'
                    title_cell.font = Font(name='微软雅黑', size=11, bold=True)
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    title_cell.fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
                    title_cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    # 合并标题单元格
                    worksheet.merge_cells(f'A{stats_start_row}:C{stats_start_row}')
                    
                    # 设置表头
                    headers = ['类型', '数量', '金额（$）']
                    for col_idx, header in enumerate(headers):
                        cell = worksheet.cell(row=stats_start_row + 1, column=col_idx + 1)
                        cell.value = header
                        cell.font = Font(name='微软雅黑', size=10, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                    
                    # 写入NAT VPS数据
                    nat_row = stats_start_row + 2
                    for col_idx, value in enumerate(['使用NAT的VPS', nat_vps_count, nat_vps_cost]):
                        cell = worksheet.cell(row=nat_row, column=col_idx + 1)
                        cell.value = value
                        cell.font = Font(name='微软雅黑', size=10, color='800080')  # 紫色
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        if col_idx == 2:  # 金额列使用货币格式
                            cell.number_format = '$#,##0.00'
                    
                    # 写入非NAT VPS数据
                    non_nat_row = stats_start_row + 3
                    for col_idx, value in enumerate(['未使用NAT的VPS', non_nat_vps_count, non_nat_vps_cost]):
                        cell = worksheet.cell(row=non_nat_row, column=col_idx + 1)
                        cell.value = value
                        cell.font = Font(name='微软雅黑', size=10, color='0000FF')  # 蓝色
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        if col_idx == 2:  # 金额列使用货币格式
                            cell.number_format = '$#,##0.00'
                    
                    # 写入NAT费用行
                    nat_fee_row = stats_start_row + 4
                    for col_idx, value in enumerate(['NAT费用', '-', nat_fee]):
                        cell = worksheet.cell(row=nat_fee_row, column=col_idx + 1)
                        cell.value = value
                        cell.font = Font(name='微软雅黑', size=10, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        if col_idx == 2:  # 金额列使用货币格式
                            cell.number_format = '$#,##0.00'
                    
                    # 写入总计行
                    total_row = stats_start_row + 5
                    for col_idx, value in enumerate(['总计', nat_vps_count + non_nat_vps_count, total_amount]):
                        cell = worksheet.cell(row=total_row, column=col_idx + 1)
                        cell.value = value
                        cell.font = Font(name='微软雅黑', size=10, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        if col_idx == 2:  # 金额列使用货币格式
                            cell.number_format = '$#,##0.00'
                    
                    # 设置列宽
                    worksheet.column_dimensions['A'].width = 20  # 类型列宽
                    worksheet.column_dimensions['B'].width = 10  # 数量列宽
                    worksheet.column_dimensions['C'].width = 15  # 金额列宽
                
                logger.info(f"成功导出{specific_year}年{specific_month}月账单到 {output_file}")
                return True
            
            # 不是指定月份导出，而是导出历史账单汇总
            else:
                # 获取当前日期
                now = datetime.datetime.now()
                end_year = now.year
                end_month = now.month
                
                # 如果指定了年份，则只导出该年份的数据
                if specific_year is not None:
                    start_year = specific_year
                    end_year = specific_year
                
                # 获取月账单表格数据
                summary_df, bill_data = self.generate_monthly_bill_table(start_year, end_year, end_month)
                
                # 检查是否有数据
                if summary_df.empty or not bill_data:
                    logger.warning(f"没有找到从{start_year}年到{end_year}年{end_month}月的账单数据")
                    return False
                
                # 确保目录存在
                output_dir = os.path.dirname(output_file)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir)
                
                # 创建Excel工作簿，使用openpyxl引擎支持更多格式设置
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    # 首先保存汇总表
                    summary_df.to_excel(writer, sheet_name='月度账单汇总', index=False)
                    
                    # 获取汇总工作表
                    summary_sheet = writer.sheets['月度账单汇总']
                    
                    # 导入样式
                    from openpyxl.styles import PatternFill, Alignment, Font, Color, Border, Side
                    
                    # 设置表头样式
                    header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
                    header_font = Font(name='微软雅黑', size=11, bold=True, color='000000')
                    
                    # 设置单元格对齐方式
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 应用表头样式
                    for cell in next(summary_sheet.iter_rows(min_row=1, max_row=1)):
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                    
                    # 设置列宽
                    summary_sheet.column_dimensions['A'].width = 10  # 年份
                    summary_sheet.column_dimensions['B'].width = 10  # 月份
                    summary_sheet.column_dimensions['C'].width = 15  # 账单日期
                    summary_sheet.column_dimensions['D'].width = 10  # VPS数量
                    summary_sheet.column_dimensions['E'].width = 10  # NAT费用
                    summary_sheet.column_dimensions['F'].width = 10  # 月总费用
                    
                    # 设置数据行样式
                    for row in range(2, summary_sheet.max_row + 1):
                        for col in range(1, summary_sheet.max_column + 1):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.alignment = center_alignment
                    
                    # 然后保存每个月份的详细账单
                    for bill in bill_data:
                        year = bill['年份']
                        month = bill['月份']
                        
                        # 设置月份名称
                        month_names = {
                            1: "一月", 2: "二月", 3: "三月", 4: "四月",
                            5: "五月", 6: "六月", 7: "七月", 8: "八月",
                            9: "九月", 10: "十月", 11: "十一月", 12: "十二月"
                        }
                        month_name = month_names.get(month, str(month) + "月")
                        
                        # 创建工作表名
                        sheet_name = f"{year}年{month_name}"
                        
                        # 确保工作表名称不超过31个字符（Excel限制）
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        
                        # 获取月份数据
                        month_data = bill['详细数据']
                        
                        # 确定列：一些月份可能包含销毁时间列
                        # 检查是否有销毁时间
                        has_destroy_time = bill.get('显示销毁时间列', False)
                        
                        if has_destroy_time:
                            columns = ['VPS名称', '是否使用NAT', '使用状态', '销毁时间', '使用时长', '单价/月（$）', '合计（$）']
                        else:
                            columns = ['VPS名称', '是否使用NAT', '使用状态', '使用时长', '单价/月（$）', '合计（$）']
                        
                        # 创建DataFrame
                        month_df = pd.DataFrame(month_data, columns=columns)
                        
                        # 添加NAT费用和总计行
                        if bill['NAT费用'] > 0:
                            # 计算使用NAT的VPS和它们的总使用天数
                            nat_vps_count = sum(1 for row in bill['详细数据'] if row.get('是否使用NAT') == '是')
                            
                            # 计算NAT VPS的总使用天数
                            total_nat_days = 0
                            for row in bill['详细数据']:
                                if row.get('是否使用NAT') == '是':
                                    usage_str = row.get('使用时长', '')
                                    # 尝试解析使用时长中的天数
                                    if '天' in usage_str:
                                        try:
                                            days_part = usage_str.split('天')[0]
                                            days = int(days_part)
                                            # 如果有小时且超过12小时，天数加1
                                            if '小时' in usage_str:
                                                hours_part = usage_str.split('天')[1].split('小时')[0]
                                                hours = int(hours_part)
                                                if hours > 12:
                                                    days += 1
                                            total_nat_days += days
                                        except (ValueError, IndexError):
                                            pass
                            
                            # 获取指定年月的实时汇率
                            exchange_rate = self.get_exchange_rate(year, month)
                            exchange_rate_display = round(1 / exchange_rate, 2) if exchange_rate > 0 else 0
                            
                            nat_row = pd.Series({
                                'VPS名称': f'NAT费用({nat_vps_count}台VPS共{total_nat_days}天×1G/天×¥1/G÷当月汇率¥{exchange_rate_display}:$1)',
                                '合计（$）': bill['NAT费用']
                            })
                            month_df = pd.concat([month_df, pd.DataFrame([nat_row])], ignore_index=True)
                        
                        # 添加总计行
                        total_row = pd.Series({
                            'VPS名称': '总计',
                            '合计（$）': bill['月总费用']
                        })
                        month_df = pd.concat([month_df, pd.DataFrame([total_row])], ignore_index=True)
                        
                        # 保存到Excel
                        month_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 获取工作表列表
                    worksheets = writer.sheets
                    
                    # 导入样式
                    from openpyxl.styles import PatternFill, Alignment, Font, Color, Border, Side
                    
                    # 设置表格整体背景为草绿色边框
                    # 使用更深的草绿色作为边框
                    thin_border = Side(border_style="thin", color="2E8B57")
                    medium_border = Side(border_style="medium", color="2E8B57")
                    bordered_cell = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
                    header_border = Border(top=medium_border, left=medium_border, right=medium_border, bottom=medium_border)
                    
                    # 自动换行和居中对齐
                    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    
                    # 为每个月份工作表设置样式
                    for sheet_name, sheet in worksheets.items():
                        if sheet_name != '月度账单汇总':
                            # 设置表头样式
                            header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
                            header_font = Font(name='微软雅黑', size=11, bold=True, color='000000')
                            
                            # 应用表头样式
                            for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = center_alignment
                                cell.border = header_border
                            
                            # 设置列宽
                            sheet.column_dimensions['A'].width = 45  # VPS名称
                            if '是否使用NAT' in month_df.columns:
                                sheet.column_dimensions['B'].width = 15  # 是否使用NAT
                                sheet.column_dimensions['C'].width = 10  # 使用状态
                                col_offset = 1
                                if has_destroy_time:
                                    sheet.column_dimensions['D'].width = 15  # 销毁时间
                                    col_offset += 1
                                sheet.column_dimensions[chr(68 + col_offset)].width = 20  # 使用时长
                                sheet.column_dimensions[chr(69 + col_offset)].width = 15  # 单价/月
                                sheet.column_dimensions[chr(70 + col_offset)].width = 15  # 合计
                            
                            # 设置数据行样式
                            for row in range(2, sheet.max_row + 1):
                                # 设置每一行的样式
                                row_font = Font(name='微软雅黑', size=10)
                                
                                # NAT费用和总计行使用粗体
                                if row >= sheet.max_row - 1:
                                    row_font = Font(name='微软雅黑', size=10, bold=True)
                                
                                for col in range(1, sheet.max_column + 1):
                                    cell = sheet.cell(row=row, column=col)
                                    cell.font = row_font
                                    cell.border = bordered_cell
                                    
                                    # 根据不同状态设置不同颜色
                                    if row < sheet.max_row - 1:  # 排除NAT费用和总计行
                                        # 获取使用状态值
                                        status_col = None
                                        for idx, col_name in enumerate(month_df.columns):
                                            if col_name == '使用状态':
                                                status_col = idx + 1
                                                break
                                        
                                        nat_col = None
                                        for idx, col_name in enumerate(month_df.columns):
                                            if col_name == '是否使用NAT':
                                                nat_col = idx + 1
                                                break
                                            
                                        if status_col and nat_col:
                                            status_cell = sheet.cell(row=row, column=status_col)
                                            nat_cell = sheet.cell(row=row, column=nat_col)
                                            
                                            status = status_cell.value
                                            use_nat = nat_cell.value
                                            
                                            # 设置颜色
                                            if status == '销毁':
                                                # 销毁的VPS设为红色
                                                cell.font = Font(name='微软雅黑', size=10, color='FF0000')
                                            elif use_nat == '是':
                                                # 使用NAT的VPS设为紫色
                                                cell.font = Font(name='微软雅黑', size=10, color='800080')
                                            elif use_nat == '否':
                                                # 不使用NAT的VPS设为蓝色
                                                cell.font = Font(name='微软雅黑', size=10, color='0000FF')
                                    
                                    # NAT费用和总计行使用特殊样式
                                    if row >= sheet.max_row - 1:
                                        cell.font = Font(name='微软雅黑', size=10, bold=True, color='000000')
                                    
                                    # VPS名称列左对齐
                                    if col == 1:
                                        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                                    # 金额列右对齐
                                    elif col >= sheet.max_column - 1:
                                        cell.alignment = Alignment(wrap_text=True, horizontal='right', vertical='center')
                                    else:
                                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                
                logger.info(f"成功导出从{start_year}年到{end_year}年{end_month}月的账单汇总到 {output_file}")
                return True
                
        except Exception as e:
            logger.error(f"保存月账单统计到Excel出错: {str(e)}", exc_info=True)
            return False

    def init_sample_vps_data(self):
        """
        初始化示例VPS数据，根据用户的需求创建VPS数据
        
        Returns:
            bool: 是否成功
        """
        try:
            # 清空现有数据
            self.vps_data = []
            
            # 创建示例数据
            sample_data = [
                {
                    'name': 'VPS-1',
                    'purchase_date': '2024/12/16',
                    'start_date': '2024/12/16',
                    'use_nat': True,
                    'status': '在用',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-2',
                    'purchase_date': '2024/12/16',
                    'start_date': '2024/12/16',
                    'use_nat': True,
                    'status': '在用',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-3',
                    'purchase_date': '2024/12/16',
                    'start_date': '2024/12/16',
                    'use_nat': True,
                    'status': '在用',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-4',
                    'purchase_date': '2025/1/16',
                    'start_date': '2025/1/16',
                    'use_nat': False,
                    'status': '销毁',
                    'cancel_date': '2025/3/28',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-5',
                    'purchase_date': '2025/2/1',
                    'start_date': '2025/2/1',
                    'use_nat': False,
                    'status': '销毁',
                    'cancel_date': '2025/3/28',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-6',
                    'purchase_date': '2025/2/3',
                    'start_date': '2025/2/3',
                    'use_nat': False,
                    'status': '在用',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-7',
                    'purchase_date': '2025/3/2',
                    'start_date': '2025/3/2',
                    'use_nat': False,
                    'status': '在用',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-8',
                    'purchase_date': '2025/3/5',
                    'start_date': '2025/3/5',
                    'use_nat': False,
                    'status': '销毁',
                    'cancel_date': '2025/3/28',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-9',
                    'purchase_date': '2025/3/8',
                    'start_date': '2025/3/8',
                    'use_nat': False,
                    'status': '销毁',
                    'cancel_date': '2025/3/28',
                    'price_per_month': 20
                },
                {
                    'name': 'VPS-10',
                    'purchase_date': '2025/3/19',
                    'start_date': '2025/3/19',
                    'use_nat': False,
                    'status': '在用',
                    'price_per_month': 20
                }
            ]
            
            # 将示例数据添加到VPS数据中
            self.vps_data = sample_data
            
            # 保存数据
            result = self.save_data()
            
            # 更新价格
            self.update_prices()
            
            logger.info(f"已初始化 {len(sample_data)} 台示例VPS数据")
            return result
        except Exception as e:
            logger.error(f"初始化示例VPS数据失败: {str(e)}", exc_info=True)
            return False

    def batch_add_vps(self, vps_list):
        """
        批量添加VPS数据
        
        Args:
            vps_list (list): 包含多个VPS数据字典的列表
            
        Returns:
            dict: 包含成功添加的数量和错误信息的字典
        """
        try:
            success_count = 0
            failed_count = 0
            errors = []
            
            for vps_data in vps_list:
                vps_name = vps_data.get('name')
                
                if not vps_name:
                    errors.append("VPS数据缺少name字段")
                    failed_count += 1
                    continue
                    
                # 检查是否已存在
                if self.get_vps_by_name(vps_name):
                    errors.append(f"VPS已存在: {vps_name}")
                    failed_count += 1
                    continue
                    
                # 添加start_date字段，记录创建时间
                if 'start_date' not in vps_data:
                    vps_data['start_date'] = vps_data.get('purchase_date') or datetime.datetime.now().strftime("%Y/%m/%d")
                    
                # 添加purchase_date字段，记录购买时间
                if 'purchase_date' not in vps_data:
                    vps_data['purchase_date'] = datetime.datetime.now().strftime("%Y/%m/%d")
                    
                # 添加VPS数据
                self.vps_data.append(vps_data)
                success_count += 1
            
            # 只有在成功添加了VPS时才保存数据
            if success_count > 0:
                save_result = self.save_data()
                if save_result:
                    # 更新价格
                    self.update_prices()
                    logger.info(f"已成功批量添加 {success_count} 台VPS数据")
                else:
                    logger.error("保存VPS数据失败")
                    return {"success": False, "message": "保存VPS数据失败", "added": 0, "failed": failed_count + success_count, "errors": errors}
            
            return {
                "success": success_count > 0,
                "message": f"已添加 {success_count} 台VPS，失败 {failed_count} 台",
                "added": success_count,
                "failed": failed_count,
                "errors": errors
            }
        except Exception as e:
            logger.error(f"批量添加VPS数据失败: {str(e)}", exc_info=True)
            return {
                "success": False,
                "message": f"批量添加VPS数据出错: {str(e)}",
                "added": 0,
                "failed": len(vps_list),
                "errors": [str(e)]
            }

# 如果作为命令行脚本运行
if __name__ == "__main__":
    # 设置日志格式
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='VPS账单管理工具')
    parser.add_argument('--action', type=str, required=True, 
                        help='要执行的操作: get_current_month_bill, get_monthly_bill, get_monthly_bill_summary, save_monthly_billing_to_excel, get_all_vps, save_vps, delete_vps, init_sample_data, update_prices, batch_add_vps')
    parser.add_argument('--year', type=int, help='指定的年份')
    parser.add_argument('--month', type=int, help='指定的月份')
    parser.add_argument('--specific_year', type=int, help='导出单个月账单时指定的年份')
    parser.add_argument('--specific_month', type=int, help='导出单个月账单时指定的月份')
    parser.add_argument('--output', type=str, help='输出文件路径')
    parser.add_argument('--config', type=str, default='vps_data.yml', help='配置文件路径')
    parser.add_argument('--vps_name', type=str, help='VPS名称')
    parser.add_argument('--vps_data', type=str, help='VPS数据JSON字符串')
    parser.add_argument('--vps_list', type=str, help='批量添加的VPS数据列表JSON字符串')
    args = parser.parse_args()
    
    # 创建账单管理器实例
    billing_manager = BillingManager(config_file=args.config)
    
    try:
        # 根据action参数执行相应操作
        if args.action == 'get_current_month_bill':
            # 获取当前月账单
            result = billing_manager.get_current_month_bill()
            # 输出JSON格式结果
            print(json.dumps(result, ensure_ascii=False))
            
        elif args.action == 'get_monthly_bill':
            # 检查是否提供了年月参数
            if args.year is None or args.month is None:
                raise ValueError("获取月账单需要指定year和month参数")
            
            # 获取指定月账单
            result = billing_manager.get_monthly_bill_data(args.year, args.month)
            # 输出JSON格式结果
            print(json.dumps(result, ensure_ascii=False))
            
        elif args.action == 'get_monthly_bill_summary':
            # 获取月账单汇总
            summary_df, bill_data = billing_manager.generate_monthly_bill_table()
            
            # 将DataFrame转换为字典列表
            summary_list = []
            for _, row in summary_df.iterrows():
                summary_list.append(row.to_dict())
            
            # 输出JSON格式结果
            print(json.dumps(summary_list, ensure_ascii=False))
            
        elif args.action == 'save_monthly_billing_to_excel':
            # 导出月账单统计到Excel
            output_file = args.output or '月账单统计.xlsx'
            
            # 检查是否提供了年月参数
            if args.specific_year is not None and args.specific_month is not None:
                # 如果提供了specific_year和specific_month参数，只导出指定月份的账单
                success = billing_manager.save_monthly_billing_to_excel(
                    output_file, 
                    specific_year=args.specific_year, 
                    specific_month=args.specific_month
                )
            elif args.year is not None and args.month is not None:
                # 兼容旧参数 year 和 month
                success = billing_manager.save_monthly_billing_to_excel(
                    output_file, 
                    specific_year=args.year, 
                    specific_month=args.month
                )
            else:
                # 否则导出所有月份的账单汇总
                success = billing_manager.save_monthly_billing_to_excel(output_file)
            
            if success:
                print(f"成功保存月账单统计到 {output_file}")
                sys.exit(0)
            else:
                print(f"保存月账单统计失败", file=sys.stderr)
                sys.exit(1)
                
        elif args.action == 'get_all_vps':
            # 获取所有VPS数据
            all_vps = billing_manager.get_all_vps()
            # 输出JSON格式结果
            print(json.dumps(all_vps, ensure_ascii=False))
            
        elif args.action == 'save_vps':
            # 检查是否提供了VPS数据
            if args.vps_data is None:
                raise ValueError("保存VPS需要提供vps_data参数")
            
            # 解析VPS数据JSON字符串
            vps_data = json.loads(args.vps_data)
            
            # 保存VPS数据
            if 'name' in vps_data:
                vps_name = vps_data['name']
                existing_vps = billing_manager.get_vps_by_name(vps_name)
                
                if existing_vps:
                    # 更新已有VPS
                    success = billing_manager.update_vps(vps_name, **vps_data)
                    result = billing_manager.get_vps_by_name(vps_name)
                else:
                    # 添加新VPS
                    success = billing_manager.add_vps(vps_data)
                    result = billing_manager.get_vps_by_name(vps_name)
                
                if success and result:
                    # 更新价格
                    billing_manager.update_prices()
                    # 输出更新后的VPS数据
                    print(json.dumps(result, ensure_ascii=False))
                else:
                    print(f"保存VPS失败: {vps_name}", file=sys.stderr)
                    sys.exit(1)
            else:
                print("VPS数据缺少name字段", file=sys.stderr)
                sys.exit(1)
                
        elif args.action == 'delete_vps':
            # 检查是否提供了VPS名称
            if args.vps_name is None:
                raise ValueError("删除VPS需要提供vps_name参数")
            
            # 删除VPS
            success = billing_manager.delete_vps(args.vps_name)
            print(json.dumps({"success": success}, ensure_ascii=False))
            
        elif args.action == 'init_sample_data':
            # 初始化示例数据
            success = billing_manager.init_sample_vps_data()
            print(json.dumps({"success": success}, ensure_ascii=False))
            
        elif args.action == 'update_prices':
            # 更新VPS价格
            billing_manager.update_prices()
            print(json.dumps({"success": True}, ensure_ascii=False))
            
        elif args.action == 'batch_add_vps':
            # 检查是否提供了VPS列表数据
            if args.vps_list is None:
                raise ValueError("批量添加VPS需要提供vps_list参数")
            
            # 解析VPS列表数据JSON字符串
            vps_list = json.loads(args.vps_list)
            
            # 批量添加VPS
            result = billing_manager.batch_add_vps(vps_list)
            print(json.dumps(result, ensure_ascii=False))
            
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1) 